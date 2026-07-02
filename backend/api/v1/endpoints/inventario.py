import io
import openpyxl
import pandas as pd
from datetime import datetime, date
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

import crud
import models
import schemas
from api.deps import get_db, get_current_active_user

router = APIRouter()

@router.get("/movimientos/template")
def get_movimientos_template(current_user: models.User = Depends(get_current_active_user)):
    wb = openpyxl.Workbook()

    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_properties.tabColor = "10B981"
    ws_inst.cell(row=2, column=2, value="🛠 CÓMO CARGAR MOVIMIENTOS").font = Font(size=14, bold=True, color="10B981")
    instrucciones = [
        "1. Usa la pestaña 'Plantilla Datos'.",
        "2. El PRODUCTO_NOMBRE debe coincidir exactamente con uno existente en el sistema.",
        "3. TIPO: Usa la lista desplegable (entrada, salida, ajuste).",
        "4. CANTIDAD: Siempre positiva (el sistema deduce si es salida)."
    ]
    for i, inst in enumerate(instrucciones, 4):
        ws_inst.cell(row=i, column=2, value=inst).font = Font(size=11)
    ws_inst.column_dimensions['B'].width = 80

    ws_datos = wb.create_sheet(title="Plantilla Datos")
    headers = ["PRODUCTO_NOMBRE", "TIPO", "CANTIDAD", "COSTO_UNITARIO", "MOTIVO", "REFERENCIA", "OBSERVACION"]
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws_datos.cell(row=1, column=col_num, value=header)
        cell.fill, cell.font = header_fill, header_font
        ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22

    dv_tipo = DataValidation(type="list", formula1='"entrada,salida,ajuste"', allow_blank=False)
    ws_datos.add_data_validation(dv_tipo)
    dv_tipo.add("B2:B1000")

    ejemplos = [
        ["Cacao Tostado", "entrada", 50, 2500, "Compra inicial", "FACT-001", "Stock base"],
        ["Chocolatina 80g", "ajuste", 5, 0, "Dañado", "MERMA", "Se rompió empaque"]
    ]
    for r_idx, row in enumerate(ejemplos, 2):
        for c_idx, val in enumerate(row, 1):
            ws_datos.cell(row=r_idx, column=c_idx, value=val)

    ws_datos.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_movimientos_PRO.xlsx"'}
    )

@router.get("/kardex/{producto_id}", response_model=schemas.KardexResponse)
def kardex_producto(
    producto_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    return crud.get_kardex_promedio_ponderado(db, empresa_id=current_user.empresa_id, producto_id=producto_id, start_date=start_date, end_date=end_date)

@router.get("/kardex/{producto_id}/export")
def kardex_export_excel(
    producto_id: int,
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    rep = crud.get_kardex_promedio_ponderado(db, empresa_id=current_user.empresa_id, producto_id=producto_id, start_date=start_date, end_date=end_date)

    data = []
    for it in rep.items:
        data.append({
            "Fecha": it.fecha.strftime("%Y-%m-%d %H:%M") if it.fecha else "",
            "Tipo": it.tipo.capitalize(),
            "Cantidad": it.cantidad,
            "Costo Unit.": it.costo_unitario,
            "Referencia": it.referencia or "",
            "Saldo Cant.": it.saldo_cantidad,
            "Saldo Costo": it.saldo_costo_unitario,
            "Saldo Valor": it.saldo_valor
        })

    df = pd.DataFrame(data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Kardex")

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="kardex_producto_{producto_id}.xlsx"',
        },
    )
@router.post("/movimientos", response_model=schemas.InventoryMovementOut)
def crear_movimiento(
    payload: schemas.InventoryMovementCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    try:
        payload.usuario_id = current_user.id
        # descontar_lotes: los movimientos manuales de salida sobre productos
        # perecederos también descuentan de sus lotes (FEFO) para no
        # desincronizar el módulo de Lotes.
        return crud.create_movement(db, empresa_id=current_user.empresa_id, payload=payload,
                                    descontar_lotes=True)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/movimientos", response_model=List[schemas.InventoryMovementOut])
def listar_movimientos(
    producto_id: Optional[int] = None,
    lote_id: Optional[int] = None,
    numero_lote: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    return crud.list_movements(db, empresa_id=current_user.empresa_id, producto_id=producto_id,
                               limit=limit, lote_id=lote_id, numero_lote=numero_lote)


@router.get("/lotes/{lote_id}/trazabilidad")
def trazabilidad_lote(
    lote_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    """Libro completo del lote para auditorías y recall: movimientos
    (entrada/salidas/ajustes) y las ventas/clientes a los que se despachó."""
    from crud.perecederos import get_trazabilidad_lote
    return get_trazabilidad_lote(db, empresa_id=current_user.empresa_id, lote_id=lote_id)

@router.get("/alertas/bajo-stock", response_model=List[schemas.InventoryAlertOut])
def alertas_bajo_stock(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    prods = crud.get_low_stock(db, empresa_id=current_user.empresa_id)
    return [
        schemas.InventoryAlertOut(producto_id=p.id, nombre=p.nombre, stock_actual=p.stock_actual or 0, stock_minimo=p.stock_minimo or 0)
        for p in prods
    ]

@router.post("/movimientos/upload", response_model=schemas.BulkLoadResponse)
def upload_movimientos(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.bulk_create_movimientos(db=db, empresa_id=current_user.empresa_id, file=file.file, filename=file.filename)

@router.post("/lotes", response_model=schemas.LoteExistenciaOut)
def crear_lote(
    payload: schemas.LoteExistenciaCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Registra un nuevo lote de existencias.
    Si el número de lote ya existe para ese producto, suma la cantidad.
    """
    lote = crud.crear_lote_existencia(db, empresa_id=current_user.empresa_id, payload=payload)
    return crud._enriquecer_lote(lote)

@router.get("/lotes", response_model=List[schemas.LoteExistenciaOut])
def listar_todos_lotes(
    solo_activos: bool   = Query(True),
    producto_id:  Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Lista todos los lotes activos de la empresa, opcionalmente filtrados por producto."""
    return crud.get_todos_los_lotes(
        db, empresa_id=current_user.empresa_id,
        solo_activos=solo_activos, producto_id=producto_id,
    )

@router.get("/lotes/{producto_id}", response_model=List[schemas.LoteExistenciaOut])
def listar_lotes_producto(
    producto_id: int,
    solo_activos: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Lista los lotes de un producto ordenados FEFO (primero el que vence antes)."""
    return crud.get_lotes_producto(
        db, empresa_id=current_user.empresa_id,
        producto_id=producto_id, solo_activos=solo_activos,
    )

@router.patch("/lotes/{lote_id}/ajuste")
def ajustar_lote(
    lote_id: int,
    ajuste: schemas.LoteAjusteCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Ajusta manualmente la cantidad de un lote.
    Positivo = entrada, negativo = salida.
    """
    lote = crud.ajustar_lote(db, empresa_id=current_user.empresa_id,
                              lote_id=lote_id, ajuste=ajuste)
    return crud._enriquecer_lote(lote)

@router.get("/lotes/{producto_id}/sugerencia-fefo")
def sugerencia_fefo(
    producto_id:        int,
    cantidad_requerida: float = Query(..., gt=0),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Devuelve qué lotes se consumirían al vender `cantidad_requerida` unidades.
    """
    return crud.sugerencia_fefo(
        db, empresa_id=current_user.empresa_id,
        producto_id=producto_id, cantidad_requerida=cantidad_requerida,
    )
