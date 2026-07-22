import io
import openpyxl
from typing import List
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import crud
import models
import schemas
from api.deps import get_db, get_current_active_user

router = APIRouter()

@router.get("/template")
def get_clientes_template(current_user: models.User = Depends(get_current_active_user)):
    wb = openpyxl.Workbook()

    # ── Hoja 1: Instrucciones ──────────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_properties.tabColor = "3B82F6"

    blue = Font(size=14, bold=True, color="3B82F6")
    ws_inst.cell(row=2, column=2, value="🛠  CÓMO USAR ESTA PLANTILLA DE TERCEROS (Clientes / Proveedores)").font = blue

    instrucciones = [
        ("PASO 1", "Ve a la pestaña 'Plantilla Datos' e ingresa tus terceros a partir de la fila 2."),
        ("PASO 2", "NO modifiques, renombres ni elimines la fila 1 (cabeceras en azul)."),
        ("NOMBRE", "Razón social o nombre completo. Obligatorio."),
        ("CEDULA", "NIT o número de documento. Obligatorio y único. Sin puntos ni guiones (ej: 9001234567). Para NIT NO incluyas el dígito verificador aquí."),
        ("TIPO_DOCUMENTO", "CC (Cédula) · NIT (Empresa/RUT) · CE (Cédula Extranjería) · PA (Pasaporte) · TI (Tarjeta Identidad). Por defecto: CC."),
        ("DV", "Dígito verificador — SOLO para NIT. Déjalo vacío para CC, CE, PA, TI. El sistema lo calcula si lo dejas en blanco para NIT."),
        ("TIPO_PERSONA", "NATURAL (persona natural/independiente) · JURIDICA (empresa, S.A.S., LTDA, etc.). Afecta el régimen tributario para Factura Electrónica."),
        ("TELEFONO", "Solo números, sin espacios ni guiones (ej: 3001234567)."),
        ("DIRECCION", "Dirección fiscal del tercero. Opcional."),
        ("FECHA_NACIMIENTO", "Fecha de cumpleaños, formato AAAA-MM-DD (ej: 1990-05-20). Opcional."),
        ("EMAIL", "⚠ Requerido para emitir Factura Electrónica (FE) a este cliente. Déjalo vacío si no emites FE."),
        ("CUPO_CREDITO", "Límite de crédito en COP. Solo números (ej: 500000). Deja 0 si no maneja crédito."),
        ("ES_CLIENTE", "SI → le vendes · NO → no le vendes."),
        ("ES_PROVEEDOR", "SI → le compras · NO → no le compras. Pueden ser ambos SI."),
        ("NOTA FE", "Para que la Factura Electrónica funcione el tercero debe tener: CEDULA/NIT, EMAIL, TIPO_DOCUMENTO y TIPO_PERSONA correctos. Persona JURIDICA usa régimen IVA responsable; NATURAL usa no responsable."),
    ]
    ws_inst.cell(row=4, column=2, value="COLUMNA").font = Font(bold=True, size=11)
    ws_inst.cell(row=4, column=3, value="DESCRIPCIÓN").font = Font(bold=True, size=11)
    for i, (col_key, desc) in enumerate(instrucciones, 5):
        ws_inst.cell(row=i, column=2, value=col_key).font = Font(bold=True, size=10, color="3B82F6")
        ws_inst.cell(row=i, column=3, value=desc).font = Font(size=10)
    ws_inst.column_dimensions['B'].width = 18
    ws_inst.column_dimensions['C'].width = 90

    # ── Hoja 2: Plantilla Datos ────────────────────────────────────────────────
    ws_datos = wb.create_sheet(title="Plantilla Datos")

    headers = [
        "nombre",         # A — obligatorio
        "cedula",         # B — obligatorio
        "tipo_documento", # C — CC/NIT/CE/PA/TI
        "dv",             # D — dígito verificador (solo NIT)
        "tipo_persona",   # E — NATURAL/JURIDICA
        "telefono",       # F
        "direccion",      # G
        "fecha_nacimiento", # H — AAAA-MM-DD
        "email",          # I — para FE
        "cupo_credito",   # J
        "es_cliente",     # K — SI/NO
        "es_proveedor",   # L — SI/NO
    ]
    col_widths = [30, 18, 16, 8, 14, 16, 30, 16, 30, 16, 14, 14]

    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_datos.cell(row=1, column=col_num, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    ws_datos.row_dimensions[1].height = 22

    # Desplegables
    dv_tipodoc = DataValidation(type="list", formula1='"CC,NIT,CE,PA,TI"', allow_blank=True)
    ws_datos.add_data_validation(dv_tipodoc)
    dv_tipodoc.add("C2:C5000")

    dv_persona = DataValidation(type="list", formula1='"NATURAL,JURIDICA"', allow_blank=True)
    ws_datos.add_data_validation(dv_persona)
    dv_persona.add("E2:E5000")

    dv_sino = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
    ws_datos.add_data_validation(dv_sino)
    dv_sino.add("K2:K5000")
    dv_sino.add("L2:L5000")

    # Datos de ejemplo
    #          nombre                    cedula        tipodoc  dv  persona    tel          direccion               fecha_nac     email                 cupo       cli   pro
    ejemplos = [
        ["Distribuidora XYZ S.A.S", "9001234567",  "NIT",  "1", "JURIDICA",  "6014445566", "Calle 10 # 5-20 Bogotá", "",           "contacto@xyz.com",  5000000,  "NO", "SI"],
        ["Juan Pérez García",        "10203040",    "CC",   "",  "NATURAL",   "3001234567", "Carrera 5 # 10-30",      "1990-05-20", "juan@gmail.com",    0,        "SI", "NO"],
        ["María López",              "52100200",    "CC",   "",  "NATURAL",   "3109876543", "",                        "",           "",                  1000000,  "SI", "NO"],
    ]

    example_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    for r_idx, row_data in enumerate(ejemplos, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_datos.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = example_fill

    ws_datos.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_terceros_PRO.xlsx"'}
    )

@router.post("/", response_model=schemas.Cliente)
def create_cliente(cliente: schemas.ClienteCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.create_cliente(db=db, empresa_id=current_user.empresa_id, cliente=cliente)

@router.post("/upload", response_model=schemas.BulkLoadResponse)
def upload_clientes(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.bulk_create_clientes(db=db, empresa_id=current_user.empresa_id, file=file.file, filename=file.filename)

@router.get("/", response_model=List[schemas.Cliente])
def read_clientes(skip: int = 0, limit: int = 500, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_clientes(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)

@router.get("/{cliente_id}", response_model=schemas.Cliente)
def read_cliente(cliente_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_cliente = crud.get_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return db_cliente

@router.put("/{cliente_id}", response_model=schemas.Cliente)
def update_cliente(cliente_id: int, cliente: schemas.ClienteCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_cliente = crud.update_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id, cliente=cliente)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return db_cliente

@router.get("/{cliente_id}/details", response_model=schemas.ClienteDetails)
def get_cliente_details(cliente_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_cliente = crud.get_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    deuda_actual = crud.get_cliente_deuda(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    return schemas.ClienteDetails(**db_cliente.__dict__, deuda_actual=deuda_actual)

@router.delete("/{cliente_id}")
def delete_cliente(cliente_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_cliente = crud.get_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Tercero no encontrado")
    bloqueos = crud.check_can_delete_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if bloqueos:
        raise HTTPException(
            status_code=409,
            detail=(
                f"No se puede eliminar '{db_cliente.nombre}' porque tiene: "
                + ", ".join(bloqueos) + ". Desactívelo en lugar de eliminarlo."
            )
        )
    crud.delete_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    return {"message": f"Tercero '{db_cliente.nombre}' eliminado correctamente"}

@router.get("/{cliente_id}/history", response_model=schemas.ClienteHistory)
def get_cliente_history(cliente_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    history = crud.get_cliente_history(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if history is None:
        raise HTTPException(status_code=404, detail="Historial no encontrado")
    return history
