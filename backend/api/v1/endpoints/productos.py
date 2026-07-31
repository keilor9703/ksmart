import io
import httpx
import openpyxl
import pandas as pd
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

import crud
import models
import schemas
from api.deps import get_db, get_current_active_user



import httpx

from fastapi import APIRouter, Depends
from sqlalchemy.orm import Session
from typing import Optional


router = APIRouter()

# ─── FUNCIONES DE BÚSQUEDA EN APIs EXTERNAS ─────────────────────────────────────

import logging
_barcode_logger = logging.getLogger("productos.barcode")

# Open*Facts bloquea con 403 a clientes sin User-Agent identificable
# (política anti-abuso). Identificarse es obligatorio para usar su API.
_LOOKUP_HEADERS = {"User-Agent": "Ksmart360/1.0 (ERP Colombia; https://ksmart360.com)"}


async def fetch_openfoodfacts(client: httpx.AsyncClient, barcode: str):
    """Busca en la base de datos de alimentos"""
    try:
        response = await client.get(
            f"https://world.openfoodfacts.org/api/v0/product/{barcode}.json",
            headers=_LOOKUP_HEADERS,
        )
        _barcode_logger.info("openfoodfacts %s -> HTTP %s", barcode, response.status_code)
        if response.status_code == 200:
            data = response.json()
            if data.get("status") == 1:
                p = data.get("product", {})
                return {
                    "nombre": p.get("product_name") or p.get("generic_name", ""),
                    "descripcion": p.get("brands", "")
                }
    except Exception as e:
        _barcode_logger.warning("openfoodfacts %s -> error: %s", barcode, e)
    return None

async def fetch_upcitemdb(client: httpx.AsyncClient, barcode: str):
    """Busca en una de las bases de datos de retail más grandes (Electrónica, ropa, etc.)"""
    try:
        # Nota: La API gratuita de UPCitemdb permite ~100 peticiones diarias por IP
        response = await client.get(
            f"https://api.upcitemdb.com/prod/trial/lookup?upc={barcode}",
            headers=_LOOKUP_HEADERS,
        )
        _barcode_logger.info("upcitemdb %s -> HTTP %s", barcode, response.status_code)
        if response.status_code == 200:
            data = response.json()
            items = data.get("items", [])
            if items:
                item = items[0]
                return {
                    "nombre": item.get("title", ""),
                    "descripcion": item.get("brand", "") or item.get("description", "")
                }
    except Exception as e:
        _barcode_logger.warning("upcitemdb %s -> error: %s", barcode, e)
    return None

async def fetch_openfacts_siblings(client: httpx.AsyncClient, barcode: str):
    """Bases hermanas de OFF: Belleza (cosméticos/aseo), Productos generales y Mascotas."""
    urls = [
        f"https://world.openbeautyfacts.org/api/v0/product/{barcode}.json",
        f"https://world.openproductsfacts.org/api/v0/product/{barcode}.json",
        f"https://world.openpetfoodfacts.org/api/v0/product/{barcode}.json",
    ]
    for url in urls:
        try:
            response = await client.get(url, headers=_LOOKUP_HEADERS)
            _barcode_logger.info("%s %s -> HTTP %s", url.split("/")[2], barcode, response.status_code)
            if response.status_code == 200:
                data = response.json()
                if data.get("status") == 1:
                    p = data.get("product", {})
                    if p.get("product_name"):
                        return {
                            "nombre": p.get("product_name", ""),
                            "descripcion": p.get("brands", "")
                        }
        except Exception as e:
            _barcode_logger.warning("%s %s -> error: %s", url.split("/")[2], barcode, e)
            continue
    return None

async def fetch_barcode_monster(client: httpx.AsyncClient, barcode: str):
    """barcode.monster — catálogo crowdsourced gratuito, sin API key."""
    try:
        response = await client.get(f"https://barcode.monster/api/{barcode}", headers=_LOOKUP_HEADERS)
        _barcode_logger.info("barcode.monster %s -> HTTP %s", barcode, response.status_code)
        if response.status_code == 200:
            data = response.json()
            desc = (data.get("description") or "").replace("(from barcode.monster)", "").strip()
            if data.get("status") == "active" and desc:
                return {"nombre": desc, "descripcion": data.get("company", "") or ""}
    except Exception as e:
        _barcode_logger.warning("barcode.monster %s -> error: %s", barcode, e)
    return None

async def fetch_web_search(client: httpx.AsyncClient, barcode: str):
    """Último recurso: buscar el EAN en la web (DuckDuckGo HTML, gratuito) y
    tomar el título del primer resultado — el mismo truco que usan las apps
    gratuitas de escaneo para productos locales que no están en los catálogos.

    ⚠️ A diferencia de los fetch_* de arriba, esto NO es una consulta por
    código de barras a una base de datos real — es una búsqueda de texto
    libre que puede devolver el nombre de una página totalmente no
    relacionada (un foro, un producto distinto que menciona el mismo número,
    spam SEO). Por eso el resultado se marca con fuente="web_no_verificado"
    en vez de tratarse como un match real — el frontend debe advertir
    claramente al usuario que revise/confirme el nombre antes de guardar.
    """
    import re, html as _html
    try:
        response = await client.get(
            "https://html.duckduckgo.com/html/",
            params={"q": barcode},
            headers={**_LOOKUP_HEADERS, "Accept": "text/html"},
        )
        _barcode_logger.info("websearch %s -> HTTP %s", barcode, response.status_code)
        if response.status_code != 200:
            return None
        titles = re.findall(r'class="result__a"[^>]*>(.*?)</a>', response.text, re.S)
        for raw in titles[:5]:
            title = _html.unescape(re.sub(r"<[^>]+>", "", raw)).strip()
            # Descartar títulos que son solo el código o demasiado genéricos
            limpio = title.replace(barcode, "").strip(" -|–·:")
            if len(limpio) >= 8:
                # Quitar sufijo de sitio ("… - Locatel Colombia", "… | Éxito")
                limpio = re.split(r"\s+[|·]\s+", limpio)[0].strip()
                return {
                    "nombre": limpio[:150],
                    "descripcion": "Encontrado por búsqueda web — no verificado, revisa el nombre",
                    "fuente": "web_no_verificado",
                }
    except Exception as e:
        _barcode_logger.warning("websearch %s -> error: %s", barcode, e)
    return None

# ─── ENDPOINT PRINCIPAL ─────────────────────────────────────────────────────────

@router.get("/sku-preview")
def preview_sku(
    grupo_id: int = Query(...),
    nombre: str = Query(...),
    atributos: Optional[str] = Query(None),  # JSON string
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Returns what the auto-generated SKU would look like without saving."""
    import json
    attrs = {}
    if atributos:
        try:
            attrs = json.loads(atributos)
        except Exception:
            pass
    from crud.productos import _generate_smart_sku, sku_exists
    return {"sku": _generate_smart_sku(db, current_user.empresa_id, grupo_id, nombre, variante_attrs=attrs or None)}


@router.get("/sku/{sku}", response_model=Optional[schemas.Producto])
def get_producto_por_sku(
    sku: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Busca un producto por su SKU interno (exacto, solo en la empresa del usuario)."""
    return db.query(models.Producto).filter(
        models.Producto.sku == sku.upper(),
        models.Producto.empresa_id == current_user.empresa_id,
        models.Producto.vigente == True,
    ).first()


@router.post("/{producto_id}/variantes", response_model=schemas.ProductoVarianteOut)
def crear_variante(
    producto_id: int,
    payload: schemas.ProductoVarianteCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    try:
        return crud.create_variante(db, empresa_id=current_user.empresa_id, producto_id=producto_id, payload=payload)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/{producto_id}/variantes", response_model=List[schemas.ProductoVarianteOut])
def listar_variantes(
    producto_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    return crud.list_variantes(db, empresa_id=current_user.empresa_id, producto_id=producto_id)


@router.put("/{producto_id}/variantes/{variante_id}", response_model=schemas.ProductoVarianteOut)
def actualizar_variante(
    producto_id: int,
    variante_id: int,
    payload: schemas.ProductoVarianteUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    try:
        return crud.update_variante(db, empresa_id=current_user.empresa_id, variante_id=variante_id, payload=payload)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.delete("/{producto_id}/variantes/{variante_id}")
def eliminar_variante(
    producto_id: int,
    variante_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    crud.delete_variante(db, empresa_id=current_user.empresa_id, variante_id=variante_id)
    return {"ok": True}


@router.get("/barcode/{barcode}", response_model=Optional[schemas.Producto])
async def get_producto_por_barcode(
    barcode: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    import asyncio
    # 1. Buscar SOLO en la empresa del usuario autenticado (sin cross-tenant)
    local_prod = db.query(models.Producto).filter(
        models.Producto.codigo_barras == barcode,
        models.Producto.empresa_id == current_user.empresa_id,
        models.Producto.vigente == True,
    ).first()

    if local_prod:
        return local_prod

    # 2. BÚSQUEDA PARALELA EN APIs PÚBLICAS (no se comparten datos entre tenants).
    # Los resultados se evalúan en orden de confiabilidad: catálogos
    # estructurados primero, búsqueda web (sin verificar) como último recurso.
    async with httpx.AsyncClient(timeout=6.0, follow_redirects=True) as client:
        resultados_catalogo = await asyncio.gather(
            fetch_openfoodfacts(client, barcode),
            fetch_openfacts_siblings(client, barcode),
            fetch_upcitemdb(client, barcode),
            fetch_barcode_monster(client, barcode),
            return_exceptions=True,
        )
        for resultado_api in resultados_catalogo:
            if resultado_api and not isinstance(resultado_api, Exception) and resultado_api.get("nombre"):
                return {
                    "id": 0,
                    "nombre": resultado_api["nombre"],
                    "codigo_barras": barcode,
                    "unidad_medida": "UND",
                    "grupo_item": 2,
                    "precio": 0.0,
                    "costo": 0.0,
                    "descripcion": resultado_api.get("descripcion", ""),
                    "empresa_id": current_user.empresa_id,
                    "fuente": "catalogo",
                }

        # Ningún catálogo estructurado lo reconoció — último recurso, marcado
        # explícitamente como no verificado (ver docstring de fetch_web_search).
        resultado_web = await fetch_web_search(client, barcode)
        if resultado_web and resultado_web.get("nombre"):
            return {
                "id": 0,
                "nombre": resultado_web["nombre"],
                "codigo_barras": barcode,
                "unidad_medida": "UND",
                "grupo_item": 2,
                "precio": 0.0,
                "costo": 0.0,
                "descripcion": resultado_web.get("descripcion", ""),
                "empresa_id": current_user.empresa_id,
                "fuente": resultado_web.get("fuente", "web_no_verificado"),
            }

    return None

@router.get("/template")
def get_productos_template(current_user: models.User = Depends(get_current_active_user)):
    wb = openpyxl.Workbook()

    # ── Hoja 1: Instrucciones ──────────────────────────────────────────────────
    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_properties.tabColor = "8B5CF6"

    purple = Font(size=14, bold=True, color="8B5CF6")
    ws_inst.cell(row=2, column=2, value="🛠  CÓMO USAR ESTA PLANTILLA DE PRODUCTOS").font = purple

    instrucciones = [
        ("PASO 1", "Ve a la pestaña 'Plantilla Datos' e ingresa tus productos a partir de la fila 2."),
        ("PASO 2", "NO modifiques, renombres ni elimines la fila 1 (cabeceras en morado)."),
        ("PASO 3", "GRUPO_ITEM — usa el desplegable o escribe el código:  MP (Materia Prima),  PT (Prod. Terminado),  AF (Activo Fijo),  INS (Insumo)."),
        ("PASO 4", "ES_SERVICIO — 0 = Producto Físico (controla inventario),  1 = Servicio/Intangible (sin inventario)."),
        ("PASO 5", "DURACION_MINUTOS — SOLO para servicios (ES_SERVICIO=1). Si pones la duración (ej: 30, 60), el servicio quedará habilitado para AGENDAMIENTO de citas. Déjalo vacío si el servicio no se agenda."),
        ("PASO 6", "UNIDAD_MEDIDA — usa el desplegable:  UND (unidades),  KGS (kilos),  GRS (gramos),  LTS (litros),  MTS (metros),  LBS (libras)."),
        ("PASO 7", "STOCK_INICIAL — cantidad de unidades con que inicia el inventario del producto. Deja 0 si aún no hay existencias."),
        ("PASO 8", "CODIGO_BARRAS — código EAN-13 / código interno. Déjalo vacío si no tiene; debe ser único por empresa."),
        ("PASO 9", "DESCRIPCION — texto libre opcional (ingredientes, especificaciones, etc.)."),
        ("PASO 10", "Cuando el archivo esté listo, guárdalo como .xlsx y súbelo desde el módulo de Productos → Carga Masiva."),
        ("NOTA",   "Los productos ya existentes (mismo nombre) serán omitidos sin error. Las filas con nombre vacío también se saltan."),
        ("NOTA 2", "Para agendar un servicio también debes asignarle trabajadores en: Agendamiento → Configurar."),
    ]
    ws_inst.cell(row=4, column=2, value="COLUMNA").font = Font(bold=True, size=11)
    ws_inst.cell(row=4, column=3, value="DESCRIPCIÓN").font = Font(bold=True, size=11)
    for i, (col_key, desc) in enumerate(instrucciones, 5):
        ws_inst.cell(row=i, column=2, value=col_key).font = Font(bold=True, size=10, color="8B5CF6")
        ws_inst.cell(row=i, column=3, value=desc).font = Font(size=10)
    ws_inst.column_dimensions['B'].width = 16
    ws_inst.column_dimensions['C'].width = 90

    # ── Hoja 2: Plantilla Datos ────────────────────────────────────────────────
    ws_datos = wb.create_sheet(title="Plantilla Datos")

    headers = [
        "nombre",           # A — obligatorio
        "precio",           # B
        "costo",            # C
        "grupo_item",       # D — dropdown MP/PT/AF/INS
        "unidad_medida",    # E — dropdown
        "es_servicio",      # F — 0/1
        "duracion_minutos", # G — solo servicios agendables
        "stock_minimo",     # H
        "stock_inicial",    # I
        "codigo_barras",    # J
        "descripcion",      # K
    ]

    col_widths = [28, 14, 14, 16, 16, 14, 18, 14, 14, 20, 40]
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws_datos.cell(row=1, column=col_num, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    ws_datos.row_dimensions[1].height = 22

    # Validaciones desplegables
    dv_grupo = DataValidation(type="list", formula1='"MP,PT,AF,INS"', allow_blank=True)
    dv_grupo.error = 'Usa: MP, PT, AF o INS'
    ws_datos.add_data_validation(dv_grupo)
    dv_grupo.add("D2:D5000")

    dv_unidad = DataValidation(type="list", formula1='"UND,KGS,GRS,LTS,MTS,LBS"', allow_blank=True)
    ws_datos.add_data_validation(dv_unidad)
    dv_unidad.add("E2:E5000")

    dv_servicio = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws_datos.add_data_validation(dv_servicio)
    dv_servicio.add("F2:F5000")

    # Datos de ejemplo (3 filas)
    ejemplos = [
        # nombre              precio   costo   grupo  unidad  serv  dur   stk_min  stk_ini  barcode          desc
        ["Cacao Tostado",     5000,    3000,   "MP",  "KGS",  0,    None, 10,      100,     "7790123456789", "Cacao tostado natural 1 Kg"],
        ["Chocolatina 80g",   1200,    450,    "PT",  "UND",  0,    None, 50,      200,     "7791234567890", "Chocolatina de leche 80 gramos"],
        ["Corte de Cabello",  25000,   0,      "PT",  "UND",  1,    30,   0,       0,       "",              "Servicio agendable de 30 min"],
    ]

    example_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
    for r_idx, row_data in enumerate(ejemplos, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws_datos.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = example_fill

    ws_datos.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_productos_PRO.xlsx"'}
    )

@router.post("/upload", response_model=schemas.BulkLoadResponse)
def upload_productos(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.bulk_create_productos(db=db, empresa_id=current_user.empresa_id, file=file.file, filename=file.filename)

@router.post("/{producto_id}/duplicate", response_model=schemas.Producto)
def duplicate_producto(
    producto_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    original = crud.get_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id)
    if not original:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    import copy
    nuevo = schemas.ProductoCreate(
        nombre=f"Copia de {original.nombre}",
        precio=original.precio,
        costo=original.costo,
        es_servicio=original.es_servicio,
        unidad_medida=original.unidad_medida,
        grupo_item=original.grupo_item,
        stock_minimo=original.stock_minimo,
        maneja_lotes=original.maneja_lotes,
        descripcion=original.descripcion,
        mostrar_en_catalogo=False,
        imagenes=list(original.imagenes or []),
        # SKU y barcode no se copian para evitar duplicados
    )
    return crud.create_producto(db=db, empresa_id=current_user.empresa_id, producto=nuevo)


@router.post("/", response_model=schemas.Producto)
def create_producto(producto: schemas.ProductoCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    # Validar barcode único en la empresa
    if producto.codigo_barras:
        existing = db.query(models.Producto).filter(
            models.Producto.empresa_id == current_user.empresa_id,
            models.Producto.codigo_barras == producto.codigo_barras,
            models.Producto.vigente == True,
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail=f"Ya existe un producto con el código de barras '{producto.codigo_barras}'.")
    return crud.create_producto(db=db, empresa_id=current_user.empresa_id, producto=producto)

@router.get("/low-stock", response_model=List[schemas.Producto])
def get_low_stock(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Productos físicos con stock_actual < stock_minimo."""
    from sqlalchemy import and_
    items = db.query(models.Producto).filter(
        models.Producto.empresa_id == current_user.empresa_id,
        models.Producto.vigente == True,
        models.Producto.es_servicio == False,
        models.Producto.stock_actual < models.Producto.stock_minimo,
    ).order_by(models.Producto.stock_actual.asc()).all()
    from crud.impuestos import attach_impuestos_to_productos
    return attach_impuestos_to_productos(db, current_user.empresa_id, items)


@router.get("/", response_model=List[schemas.Producto])
def read_productos(
    skip: int = 0,
    limit: int = 500,
    solo_pos: bool = False,
    q: Optional[str] = Query(None, description="Buscar por nombre, SKU, barcode, categoría"),
    grupo: Optional[str] = Query(None, description="ID de grupo o 'servicio'"),
    stock: Optional[str] = Query(None, description="ok | bajo | sinStock"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    return crud.get_productos(
        db,
        empresa_id=current_user.empresa_id,
        skip=skip,
        limit=limit,
        solo_pos=solo_pos,
        search=q,
        filter_group=grupo,
        filter_stock=stock,
    )

@router.put("/{producto_id}", response_model=schemas.Producto)
def update_producto(producto_id: int, producto: schemas.ProductoCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if producto.codigo_barras:
        existing = db.query(models.Producto).filter(
            models.Producto.empresa_id == current_user.empresa_id,
            models.Producto.codigo_barras == producto.codigo_barras,
            models.Producto.vigente == True,
            models.Producto.id != producto_id,
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail=f"Ya existe otro producto con el código de barras '{producto.codigo_barras}'.")
    db_producto = crud.update_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id, producto=producto)
    if db_producto is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return db_producto

@router.delete("/{producto_id}")
def delete_producto(producto_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_producto = crud.get_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id)
    if db_producto is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    crud.delete_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id)
    return {"message": f"Producto '{db_producto.nombre}' eliminado correctamente"}

@router.get("/export")
def exportar_productos(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    prods = crud.get_productos(db, empresa_id=current_user.empresa_id)
    groups = {1: 'MP', 2: 'PT', 3: 'AF', 4: 'INS'}
    rows = [
        {
            "id": p.id, "nombre": p.nombre, "precio": p.precio, "costo": p.costo,
            "grupo_item": groups.get(p.grupo_item, 'PT'), "es_servicio": "SÍ" if p.es_servicio else "NO",
            "unidad_medida": p.unidad_medida, "stock_minimo": float(p.stock_minimo or 0),
            "stock_actual": float(p.stock_actual or 0)
        }
        for p in prods
    ]
    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Productos")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="productos_existentes.xlsx"'}
    )

@router.patch("/{producto_id}/stock-minimo")
def actualizar_stock_minimo(
    producto_id: int,
    body: schemas.ProductoStockUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    prod = crud.update_producto_stock_minimo(db, empresa_id=current_user.empresa_id, producto_id=producto_id, minimo=body.stock_minimo or 0)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return {"ok": True}
