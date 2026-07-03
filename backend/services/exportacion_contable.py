"""
Exportación de libros contables a Excel y PDF.
Usa openpyxl (Excel) y ReportLab (PDF) — ambos ya en requirements.txt.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import List, Optional

from sqlalchemy.orm import Session

import crud
import models
import schemas


# ─── Colores corporativos ─────────────────────────────────────────────────────
_COL_HEADER = "0F172A"
_COL_ACENTO = "6366F1"
_COL_FILA_PAR = "F8FAFC"


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def _wb_header(ws, cols: List[tuple], title: str):
    """Escribe la fila de encabezado y devuelve el ancho por columna."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="CBD5E1")
    border = Border(bottom=thin)
    fill = PatternFill("solid", fgColor=_COL_HEADER)
    for col_idx, (header, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22


def _money(val) -> str:
    return f"${val:,.0f}" if val else "$0"


def exportar_libro_diario_excel(
    db: Session, empresa_id: int,
    fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    items, _ = crud.listar_asientos(db, empresa_id, skip=0, limit=5000,
                                     fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    cols = [
        ("N° Asiento", 12), ("Fecha", 14), ("Tipo", 16), ("Descripción", 40),
        ("Código", 10), ("Cuenta", 36), ("Débito", 16), ("Crédito", 16),
    ]
    _wb_header(ws, cols, "Libro Diario")

    thin = Side(style="thin", color="E2E8F0")
    border = Border(bottom=Border(bottom=thin).bottom)
    fila_par = PatternFill("solid", fgColor=_COL_FILA_PAR)

    row = 2
    for asiento in items:
        for i, linea in enumerate(sorted(asiento.lineas, key=lambda x: x.orden)):
            values = [
                asiento.numero if i == 0 else "",
                asiento.fecha.strftime("%Y-%m-%d") if i == 0 else "",
                asiento.tipo_origen if i == 0 else "",
                asiento.descripcion if i == 0 else "",
                linea.cuenta.codigo if linea.cuenta else "",
                linea.cuenta.nombre if linea.cuenta else "",
                linea.debito if linea.debito else "",
                linea.credito if linea.credito else "",
            ]
            fill = fila_par if row % 2 == 0 else None
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(size=9)
                if fill:
                    cell.fill = fill
                if col in (7, 8) and val:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_balance_comprobacion_excel(
    db: Session, empresa_id: int,
    fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = crud.get_balance_comprobacion(db, empresa_id, fecha_inicio, fecha_fin)

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance de Comprobación"

    cols = [
        ("Código", 10), ("Cuenta", 38), ("Tipo", 14), ("Naturaleza", 12),
        ("Total Débitos", 16), ("Total Créditos", 16), ("Saldo Débito", 16), ("Saldo Crédito", 16),
    ]
    _wb_header(ws, cols, "Balance de Comprobación")

    fila_par = PatternFill("solid", fgColor=_COL_FILA_PAR)
    for i, r in enumerate(rows, 2):
        fill = fila_par if i % 2 == 0 else None
        values = [r.codigo, r.nombre, r.tipo, r.naturaleza,
                  r.total_debitos, r.total_creditos, r.saldo_debito, r.saldo_credito]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(size=9)
            if fill:
                cell.fill = fill
            if col >= 5 and val:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")

    # Totales
    tr = len(rows) + 2
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(bold=True, size=10)
    for col, attr in enumerate(["total_debitos", "total_creditos", "saldo_debito", "saldo_credito"], 5):
        total = sum(getattr(r, attr) for r in rows)
        cell = ws.cell(row=tr, column=col, value=total)
        cell.font = Font(bold=True, size=10)
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════════════

def _pdf_doc(title: str, empresa_nombre: str, periodo: str):
    """Devuelve (buf, doc) listos para construir el PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=1.5*cm)
    return buf, doc


def _tabla_style(header_color="#0F172A"):
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors
    hc = colors.HexColor(header_color if header_color.startswith("#") else f"#{header_color}")
    return TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), hc),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID",          (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])


def _header_paragraph(title: str, empresa: str, periodo: str, styles):
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    h = ParagraphStyle("h", parent=styles["Title"], fontSize=14, textColor=colors.HexColor(f"#{_COL_ACENTO}"))
    s = ParagraphStyle("s", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    return [
        Paragraph(empresa, h),
        Paragraph(title, ParagraphStyle("t", parent=styles["Heading2"], fontSize=11)),
        Paragraph(periodo, s),
        Spacer(1, 10),
    ]


def exportar_estado_resultados_pdf(
    db: Session, empresa_id: int,
    fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime],
    empresa_nombre: str = "Empresa",
) -> bytes:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4

    er = crud.get_estado_resultados(db, empresa_id, fecha_inicio, fecha_fin)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    periodo = f"{fecha_inicio.strftime('%d/%m/%Y') if fecha_inicio else 'Inicio'} — {fecha_fin.strftime('%d/%m/%Y') if fecha_fin else 'Hoy'}"

    def row(label, val, bold=False, indent=0, separator=False):
        indent_str = "  " * indent
        lbl = f"<b>{indent_str}{label}</b>" if bold else f"{indent_str}{label}"
        v = f"<b>{_money(val)}</b>" if bold else _money(val)
        color = "#16A34A" if val >= 0 else "#DC2626"
        return [Paragraph(lbl, styles["Normal"]), Paragraph(f'<font color="{color}">{v}</font>', ParagraphStyle("r", parent=styles["Normal"], alignment=2))]

    story = _header_paragraph("ESTADO DE RESULTADOS", empresa_nombre, periodo, styles)
    data = [
        ["CONCEPTO", "VALOR"],
        row("(+) Venta de mercancías (4135)", er.ingresos_operacionales, indent=1),
        row("(+) Ingresos por servicios (4175)", er.ingresos_servicios, indent=1),
        row("(+) Intereses y rendimientos (4210)", er.ingresos_financieros, indent=1),
        row("TOTAL INGRESOS", er.total_ingresos, bold=True),
        row("(-) Costo de mercancías vendidas (6135)", er.costo_ventas, indent=1),
        row("UTILIDAD BRUTA", er.utilidad_bruta, bold=True),
        row("(-) Gastos operacionales (5105/5195)", er.gastos_operacionales, indent=1),
        row("(-) Gastos no operacionales (5305)", er.gastos_no_operacionales, indent=1),
        row("TOTAL GASTOS", er.total_gastos, bold=True),
        row("UTILIDAD NETA DEL PERÍODO", er.utilidad_neta, bold=True),
    ]
    t = Table(data, colWidths=[12*cm, 5*cm])
    style = _tabla_style()
    style.add("ALIGN", (1, 0), (1, -1), "RIGHT")
    style.add("FONTSIZE", (0, -1), (-1, -1), 9)
    style.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#EEF2FF"))
    t.setStyle(style)
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def exportar_balance_general_pdf(
    db: Session, empresa_id: int,
    fecha_corte: Optional[datetime],
    empresa_nombre: str = "Empresa",
) -> bytes:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4, landscape

    bg = crud.get_balance_general(db, empresa_id, fecha_corte)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    fecha_str = fecha_corte.strftime("%d/%m/%Y") if fecha_corte else "Hoy"
    story = _header_paragraph("BALANCE GENERAL", empresa_nombre, f"Al {fecha_str}", styles)

    def seccion_tabla(titulo, items, total, header_color):
        rows = [[titulo, ""]]
        for item in items:
            rows.append([f"  {item['codigo']} — {item['nombre']}", _money(item['saldo'])])
        rows.append([f"TOTAL {titulo}", _money(total)])
        t = Table(rows, colWidths=[13*cm, 5*cm])
        st = _tabla_style(header_color)
        st.add("ALIGN", (1, 0), (1, -1), "RIGHT")
        st.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
        _hc = header_color.lstrip("#")
        st.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(f"#{_hc}22"))
        t.setStyle(st)
        return t

    story += [
        seccion_tabla("ACTIVOS", bg["activos"], bg["total_activos"], "1E3A8A"),
        Spacer(1, 8),
        seccion_tabla("PASIVOS", bg["pasivos"], bg["total_pasivos"], "7F1D1D"),
        Spacer(1, 4),
        seccion_tabla("PATRIMONIO", bg["patrimonio"], bg["total_patrimonio"], "4C1D95"),
        Spacer(1, 8),
    ]

    resumen = Table([
        ["TOTAL ACTIVOS", _money(bg["total_activos"])],
        ["TOTAL PASIVOS + PATRIMONIO", _money(bg["total_pasivos_patrimonio"])],
    ], colWidths=[13*cm, 5*cm])
    resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF2FF")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#6366F1")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(resumen)
    doc.build(story)
    return buf.getvalue()


def exportar_libro_diario_pdf(
    db: Session, empresa_id: int,
    fecha_inicio: Optional[datetime], fecha_fin: Optional[datetime],
    empresa_nombre: str = "Empresa",
) -> bytes:
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4, landscape

    items, _ = crud.listar_asientos(db, empresa_id, skip=0, limit=2000,
                                     fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1.5*cm, rightMargin=1.5*cm,
                             topMargin=2*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    periodo = f"{fecha_inicio.strftime('%d/%m/%Y') if fecha_inicio else 'Inicio'} — {fecha_fin.strftime('%d/%m/%Y') if fecha_fin else 'Hoy'}"
    story = _header_paragraph("LIBRO DIARIO", empresa_nombre, periodo, styles)

    rows = [["N°", "Fecha", "Tipo", "Descripción", "Cód.", "Cuenta", "Débito", "Crédito"]]
    for asiento in items:
        for i, linea in enumerate(sorted(asiento.lineas, key=lambda x: x.orden)):
            rows.append([
                str(asiento.numero) if i == 0 else "",
                asiento.fecha.strftime("%d/%m/%y") if i == 0 else "",
                asiento.tipo_origen if i == 0 else "",
                asiento.descripcion[:40] if i == 0 else "",
                linea.cuenta.codigo if linea.cuenta else "",
                (linea.cuenta.nombre[:32] if linea.cuenta else ""),
                _money(linea.debito) if linea.debito else "",
                _money(linea.credito) if linea.credito else "",
            ])

    t = Table(rows, colWidths=[1.2*cm, 1.8*cm, 2.2*cm, 6.5*cm, 1.4*cm, 6*cm, 2.5*cm, 2.5*cm])
    t.setStyle(_tabla_style())
    story.append(t)
    doc.build(story)
    return buf.getvalue()


# ─── REPORTE FISCAL PARA CONTADOR EXTERNO ─────────────────────────────────────

def exportar_reporte_fiscal_excel(db, empresa_id: int, fecha_inicio=None, fecha_fin=None) -> bytes:
    """Paquete de insumos del período para que un contador externo liquide
    retención en la fuente, ICA e IVA: auxiliar de ventas (con NIT del
    cliente y base gravable), auxiliar de compras (con NIT del proveedor),
    gastos, y un resumen fiscal con la identificación tributaria de la
    empresa (persona natural/jurídica, régimen)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import models
    from crud.contabilidad import _rango_utc
    from crud.common import BOGOTA_TZ

    utc_ini, utc_fin = _rango_utc(fecha_inicio, fecha_fin)

    empresa = db.query(models.Empresa).filter(models.Empresa.id == empresa_id).first()
    tipo_persona = "Persona Jurídica" if (getattr(empresa, "tipo_organizacion_id", 1) or 1) == 1 else "Persona Natural"

    def _local(dt):
        if not dt:
            return ""
        try:
            if dt.tzinfo is None:
                from datetime import timezone as _tz
                dt = dt.replace(tzinfo=_tz.utc)
            return dt.astimezone(BOGOTA_TZ).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(dt)

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    def _sheet(ws, headers, widths):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill, cell.font = header_fill, header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    # ── Hoja 1: Resumen fiscal ──
    ws = wb.active
    ws.title = "Resumen Fiscal"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 30
    info_rows = [
        ("REPORTE FISCAL PARA CONTADOR", ""),
        ("Empresa", empresa.nombre if empresa else ""),
        ("NIT", getattr(empresa, "nit", "") or "—"),
        ("Tipo de contribuyente", tipo_persona),
        ("Responsabilidades fiscales (RUT)", getattr(empresa, "responsabilidad_fiscal_codes", "") or "—"),
        ("Período", f"{fecha_inicio or 'inicio'} a {fecha_fin or 'hoy'}"),
        ("", ""),
    ]

    # Totales del período (mismas consultas que los módulos, con corte UTC)
    q_v = db.query(models.Venta).filter(
        models.Venta.empresa_id == empresa_id,
        models.Venta.tipo == "venta",
        models.Venta.estado_pago.in_(["pagado", "parcial"]),
    )
    if utc_ini: q_v = q_v.filter(models.Venta.fecha >= utc_ini)
    if utc_fin: q_v = q_v.filter(models.Venta.fecha <= utc_fin)
    ventas = q_v.order_by(models.Venta.fecha).all()

    q_c = db.query(models.Compra).filter(models.Compra.empresa_id == empresa_id)
    if utc_ini: q_c = q_c.filter(models.Compra.fecha >= utc_ini)
    if utc_fin: q_c = q_c.filter(models.Compra.fecha <= utc_fin)
    compras = q_c.order_by(models.Compra.fecha).all()

    q_g = db.query(models.Gasto).filter(models.Gasto.empresa_id == empresa_id)
    if utc_ini: q_g = q_g.filter(models.Gasto.fecha >= utc_ini)
    if utc_fin: q_g = q_g.filter(models.Gasto.fecha <= utc_fin)
    gastos = q_g.order_by(models.Gasto.fecha).all()

    tot_v  = sum(float(v.total or 0) for v in ventas)
    iva_v  = sum(float(v.iva_total or 0) for v in ventas)
    tot_c  = sum(float(c.total or 0) for c in compras)
    iva_c  = sum(float(c.iva_total or 0) for c in compras)
    tot_g  = sum(float(g.monto or 0) for g in gastos)

    info_rows += [
        ("INGRESOS DEL PERÍODO (ventas cobradas)", ""),
        ("  Ventas brutas (IVA incluido)", round(tot_v, 2)),
        ("  Base gravable ventas", round(tot_v - iva_v, 2)),
        ("  IVA generado (2408)", round(iva_v, 2)),
        ("", ""),
        ("COMPRAS DEL PERÍODO", ""),
        ("  Compras brutas (IVA incluido)", round(tot_c, 2)),
        ("  Base compras", round(tot_c - iva_c, 2)),
        ("  IVA descontable (1355)", round(iva_c, 2)),
        ("", ""),
        ("  Gastos del período", round(tot_g, 2)),
        ("  IVA neto estimado (generado - descontable)", round(iva_v - iva_c, 2)),
        ("", ""),
        ("NOTAS PARA EL CONTADOR", ""),
        ("  ICA", "La base son los ingresos brutos del período; aplique la tarifa por mil de la actividad y municipio."),
        ("  Retención en la fuente", "Verifique bases mínimas en UVT y calidad de agente retenedor según el tipo de contribuyente indicado arriba."),
        ("  Detalle por tercero", "Use las hojas Ventas, Compras y Gastos (incluyen NIT/cédula de cada tercero)."),
    ]
    for r, (a, b) in enumerate(info_rows, 1):
        ca = ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        if a and (a.isupper() or r == 1):
            ca.font = Font(bold=True, size=11)

    # ── Hoja 2: Ventas ──
    ws_v = wb.create_sheet("Ventas")
    _sheet(ws_v, ["FECHA (Colombia)", "N° VENTA", "FACTURA DIAN", "CLIENTE", "NIT / CÉDULA",
                  "BASE", "IVA", "TOTAL", "MÉTODO PAGO", "ORIGEN"],
           [18, 10, 16, 28, 16, 14, 12, 14, 14, 14])
    for r, v in enumerate(ventas, 2):
        base = float(v.total or 0) - float(v.iva_total or 0)
        ws_v.cell(row=r, column=1, value=_local(v.fecha))
        ws_v.cell(row=r, column=2, value=f"V{(v.numero_venta or v.id):04d}")
        ws_v.cell(row=r, column=3, value=v.numero_factura or "")
        ws_v.cell(row=r, column=4, value=v.cliente.nombre if v.cliente else "Consumidor final")
        ws_v.cell(row=r, column=5, value=(v.cliente.cedula if v.cliente else "") or "")
        ws_v.cell(row=r, column=6, value=round(base, 2))
        ws_v.cell(row=r, column=7, value=round(float(v.iva_total or 0), 2))
        ws_v.cell(row=r, column=8, value=round(float(v.total or 0), 2))
        ws_v.cell(row=r, column=9, value=v.metodo_pago or "")
        ws_v.cell(row=r, column=10, value=v.origen or "erp")

    # ── Hoja 3: Compras ──
    ws_c = wb.create_sheet("Compras")
    _sheet(ws_c, ["FECHA (Colombia)", "N° COMPRA", "PROVEEDOR", "NIT / CÉDULA",
                  "FACTURA PROV.", "BASE", "IVA", "TOTAL", "ESTADO PAGO"],
           [18, 12, 28, 16, 16, 14, 12, 14, 14])
    for r, c in enumerate(compras, 2):
        base = float(c.total or 0) - float(c.iva_total or 0)
        prov = getattr(c, "proveedor", None)
        ws_c.cell(row=r, column=1, value=_local(c.fecha))
        ws_c.cell(row=r, column=2, value=f"OC-{(c.numero_compra or c.id):04d}")
        ws_c.cell(row=r, column=3, value=prov.nombre if prov else "")
        ws_c.cell(row=r, column=4, value=(prov.cedula if prov else "") or "")
        ws_c.cell(row=r, column=5, value=getattr(c, "referencia_factura", "") or "")
        ws_c.cell(row=r, column=6, value=round(base, 2))
        ws_c.cell(row=r, column=7, value=round(float(c.iva_total or 0), 2))
        ws_c.cell(row=r, column=8, value=round(float(c.total or 0), 2))
        ws_c.cell(row=r, column=9, value=c.estado_pago or "")

    # ── Hoja 4: Gastos ──
    ws_g = wb.create_sheet("Gastos")
    _sheet(ws_g, ["FECHA (Colombia)", "TERCERO", "NIT / CÉDULA", "CONCEPTO", "CATEGORÍA", "MONTO", "MÉTODO PAGO"],
           [18, 26, 16, 34, 16, 14, 14])
    for r, g in enumerate(gastos, 2):
        terc = getattr(g, "tercero", None)
        ws_g.cell(row=r, column=1, value=_local(g.fecha))
        ws_g.cell(row=r, column=2, value=terc.nombre if terc else "")
        ws_g.cell(row=r, column=3, value=(terc.cedula if terc else "") or "")
        ws_g.cell(row=r, column=4, value=g.concepto or "")
        ws_g.cell(row=r, column=5, value=getattr(g, "categoria", "") or "")
        ws_g.cell(row=r, column=6, value=round(float(g.monto or 0), 2))
        ws_g.cell(row=r, column=7, value=getattr(g, "metodo_pago", "") or "")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
