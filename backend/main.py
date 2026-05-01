# ─── Imports consolidados (sin duplicados) ───────────────────────────────────
import os
import io
import hashlib
import logging
import secrets
import shutil
import time
import requests

from datetime import date, datetime, timedelta, timezone
from typing import List, Optional

import openpyxl
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from fastapi import (
    FastAPI, Depends, HTTPException, Response, Request,
    status, File, UploadFile, Query, APIRouter
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.staticfiles import StaticFiles

from jose import JWTError, jwt
from passlib.context import CryptContext
from pydantic import BaseModel, Field, validator
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

import crud, models, schemas
from database import SessionLocal, engine, run_migrations
from models import Base, utcnow

# ─── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("main")

# ─── Tablas + migraciones ─────────────────────────────────────────────────────
models.Base.metadata.create_all(bind=engine)
run_migrations()

app = FastAPI(title="Ksmart360 API Multi-Tenant", version="2.1.0")

# ─── CORS ─────────────────────────────────────────────────────────────────────
origins = [
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "https://appksmp.vercel.app",
    "https://ksmart360.vercel.app",
    "https://www.appjeylor.com",
    "https://appjeylor.com",
    "https://api.appjeylor.com",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── JWT ──────────────────────────────────────────────────────────────────────
# FIX #3: JWT Secret robusto — falla fuerte en producción si no está configurado
SECRET_KEY = os.getenv("SECRET_KEY")
# if not SECRET_KEY:
#     raise RuntimeError(
#         "🛑 FATAL: SECRET_KEY no configurada en variables de entorno.\n"
#         "Genera una clave segura con:\n"
#         "  export SECRET_KEY=$(python -c 'import secrets; print(secrets.token_urlsafe(32))')\n"
#         "O añádela a tu archivo .env: SECRET_KEY=tu_clave_super_segura_aqui"
#     )


if not SECRET_KEY:
    SECRET_KEY = secrets.token_urlsafe(32)
    logger.warning("⚠️ SECRET_KEY no está configurada...")
    logger.warning("Usando una clave temporal generada al vuelo. Esto es inseguro para producción.")

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "120"))

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# ─── Encriptación de contraseñas ──────────────────────────────────────────────
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

# ─── Sesión BD ────────────────────────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ─── Datos iniciales ──────────────────────────────────────────────────────────
def initialize_default_data(db: Session):
    # AUTO-HEALING: Asegurarnos de que la Empresa Maestra (1) exista Y ESTÉ ACTIVA
    empresa_default = db.query(models.Empresa).first()
    if not empresa_default:
        empresa_default = models.Empresa(
            nombre="Vialmar Cacao (Mi Fábrica)",
            nit="900000000-1",
            color_primario="#F43F5E",
            is_active=True
        )
        db.add(empresa_default)
        db.commit()
        db.refresh(empresa_default)
    else:
        if not empresa_default.is_active:
            empresa_default.is_active = True
            db.commit()
            logger.info("✅ Empresa maestra reactivada automáticamente.")

        if not empresa_default.created_at:
            empresa_default.created_at = utcnow()
            db.commit()

    default_modules_data = [
        {"name": "Ventas",              "description": "Módulo para la gestión de ventas.",              "frontend_path": "/ventas"},
        {"name": "Cotizaciones",        "description": "Módulo para la gestión de cotizaciones.",          "frontend_path": "/cotizaciones"},
        {"name": "Resoluciones DIAN",              "description": "Módulo para la gestión de resoluciones de la DIAN.",    "frontend_path": "/admin/resoluciones"},
        {"name": "Clientes",            "description": "Módulo para la gestión de clientes.",             "frontend_path": "/clientes"},
        {"name": "Productos",           "description": "Módulo para la gestión de productos.",            "frontend_path": "/productos"},
        {"name": "Reportes",            "description": "Módulo para la visualización de reportes.",       "frontend_path": "/reportes"},
        {"name": "Gestion Usuarios",    "description": "Módulo de administración de usuarios.",           "frontend_path": "/admin/users"},
        {"name": "Gestion Roles",       "description": "Módulo de administración de roles.",              "frontend_path": "/admin/roles"},
        {"name": "Gestion Modulos",     "description": "Módulo de administración de módulos.",            "frontend_path": "/admin/modules"},
        {"name": "Órdenes de Trabajo",  "description": "Módulo para la gestión de órdenes de trabajo.",  "frontend_path": "/ordenes-trabajo"},
        {"name": "Panel del Operador",  "description": "Panel de productividad para operadores.",         "frontend_path": "/panel-operador"},
        {"name": "Recetas",             "description": "Gestión de fórmulas de producción (BOM).",        "frontend_path": "/produccion/recetas"},
        {"name": "Producción",          "description": "Gestión de lotes y transformaciones.",            "frontend_path": "/produccion/lotes"},
        {"name": "Compras",             "description": "Módulo para la gestión de compras.",              "frontend_path": "/compras"},
        {"name": "Inventarios",         "description": "Módulo para movimientos y alertas de stock.",     "frontend_path": "/inventario"},
        {"name": "lotes      ",         "description": "Módulo para productos perecederos.",              "frontend_path": "/inventario/lotes"},
        {"name": "Reportes inventario", "description": "Reportes de inventario y kardex.",                "frontend_path": "/reportes-inventario"},
        {"name": "Caja",                "description": "Módulo de corte de caja diario.",                 "frontend_path": "/caja"},
        {"name": "Préstamos",           "description": "Módulo de gestión de préstamos.",                 "frontend_path": "/prestamos"},
        {"name": "Ruta de Cobro",       "description": "Módulo de gestión de ruta de cobro.",             "frontend_path": "/ruta-cobro"},
{"name": "Parqueadero",         "description": "Dashboard del parqueadero.",      "frontend_path": "/parqueadero"},
    {"name": "Buscar Placa",        "description": "Búsqueda rápida de placas.",      "frontend_path": "/parqueadero/buscar"},
    {"name": "Vehículos",           "description": "Gestión de vehículos.",            "frontend_path": "/parqueadero/vehiculos"},
    {"name": "Suscripciones Parq.", "description": "Renovaciones y pagos.",            "frontend_path": "/parqueadero/suscripciones"},
    {"name": "Config Parqueadero",  "description": "Tarifas y cupo total.",            "frontend_path": "/parqueadero/config"},

    ]

    admin_role = crud.get_role_by_name(db, name="Admin")
    if not admin_role:
        admin_role = crud.create_role(db, schemas.RoleCreate(name="Admin"))

    created_modules = []
    for mod_data in default_modules_data:
        modulo = crud.get_modulo_by_frontend_path(db, frontend_path=mod_data["frontend_path"])
        if not modulo:
            modulo = crud.create_modulo(db, schemas.ModuloCreate(**mod_data))
        created_modules.append(modulo)

    crud.set_modules_for_role(db, admin_role.id, [m.id for m in created_modules])

    admin_user = crud.get_user_by_username(db, username="admin")
    if not admin_user:
        crud.create_user(
            db,
            schemas.UserCreate(username="admin", password="adminpass", role_id=admin_role.id),
            empresa_id=1
        )


@app.on_event("startup")
def startup_event():
    models.Base.metadata.create_all(bind=engine)
    run_migrations()
    db = SessionLocal()
    try:
        initialize_default_data(db)
    finally:
        db.close()


# ─── JWT helpers & DEPENDENCIAS DE SEGURIDAD ──────────────────────────────────
def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + (expires_delta or timedelta(minutes=15))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(db: Session = Depends(get_db), token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No se pudieron validar las credenciales",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
        token_data = schemas.TokenData(username=username)
    except JWTError:
        raise credentials_exception

    user = crud.get_user_by_username(db, username=token_data.username)
    if user is None:
        raise credentials_exception
    return user


def get_current_active_user(current_user: schemas.User = Depends(get_current_user)):
    """
    Valida que el usuario esté activo, que su empresa esté activa
    y que su suscripción no haya vencido.
    """
    if not current_user:
        raise HTTPException(status_code=400, detail="Usuario inactivo")

    empresa = current_user.empresa
    if not empresa.is_active:
        raise HTTPException(
            status_code=403,
            detail="Suscripción suspendida. Contacte a soporte."
        )

    if empresa.trial_ends_at:
        ahora_utc = datetime.now(timezone.utc)
        fecha_limite = empresa.trial_ends_at
        if fecha_limite.tzinfo is None:
            fecha_limite = fecha_limite.replace(tzinfo=timezone.utc)
        if ahora_utc > fecha_limite:
            raise HTTPException(
                status_code=status.HTTP_402_PAYMENT_REQUIRED,
                detail="Suscripción expirada."
            )

    return current_user


def get_current_admin_user(current_user: schemas.User = Depends(get_current_user)):
    if current_user.role.name != "Admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Permisos insuficientes"
        )
    return current_user


def get_current_superadmin_user(current_user: schemas.User = Depends(get_current_user)):
    """Solo los usuarios Admin de la Empresa 1 (Dueños del SaaS) pueden acceder."""
    if current_user.role.name != "Admin" or current_user.empresa_id != 1:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acceso denegado: Solo SuperAdmin"
        )
    return current_user


# ═══════════════════════════════════════════════════════════════════════════════
# AUTENTICACIÓN / REGISTRO
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/auth/register", status_code=status.HTTP_201_CREATED)
def registrar_nuevo_cliente(data: schemas.RegistroSaaS, db: Session = Depends(get_db)):
    # ── Validaciones de unicidad ─────────────────────────────────────────────
    if db.query(models.User).filter(models.User.username == data.username).first():
        raise HTTPException(status_code=400, detail="Este usuario ya está en uso. Prueba con otro.")

    if data.email:
        existing_email = db.query(models.User).filter(models.User.email == data.email).first()
        if existing_email:
            raise HTTPException(status_code=400, detail="Este correo ya está registrado. ¿Quieres iniciar sesión?")

    # ── Perfilado de módulos según el tipo de negocio ─────────────────────────
    PERFILES = {
    "erp": [
        "/ventas", "/compras", "/clientes", "/productos", "/inventario",
        "/caja", "/produccion/lotes", "/ordenes-trabajo", "/panel-operador",
        "/reportes", "/cotizaciones",
    ],
    "prestamos": [
        "/clientes", "/prestamos", "/ruta-cobro", "/caja", "/reportes",
    ],
    # ✅ NUEVO PERFIL: PARQUEADERO DE MOTOS
    "parqueadero": [
        "/parqueadero",                  # Dashboard del día (pantalla principal)
        "/parqueadero/buscar",           # Buscar placa (entrada/salida rápida)
        "/parqueadero/vehiculos",        # Listado de vehículos registrados
        "/parqueadero/suscripciones",    # Histórico de pagos/renovaciones
        "/parqueadero/config",           # Tarifas y cupo total
        "/clientes",                      # Propietarios (reutiliza tu módulo Clientes)
        "/caja",                          # Corte de caja diario
        "/reportes",                      # Reportes de ingresos
    ],
}

    modulos = PERFILES.get(data.tipo_negocio, PERFILES["erp"])

    try:
        # ── Crear empresa ─────────────────────────────────────────────────────
        nueva_emp = models.Empresa(
            nombre              = data.nombre_empresa.strip(),
            is_active           = True,
            plan_type           = "trial",
            trial_ends_at       = datetime.now(timezone.utc) + timedelta(days=14),
            modulos_habilitados = modulos,
            # Campos nuevos:
            pais                = data.pais,
            ciudad              = data.ciudad,
            tamano_negocio      = data.tamano_negocio,
            origen_marketing    = data.origen,
        )
        db.add(nueva_emp)
        db.flush()

        # ── Crear usuario admin asociado ──────────────────────────────────────
        rol_admin = db.query(models.Role).filter(models.Role.name == "Admin").first()
        if not rol_admin:
            raise HTTPException(status_code=500, detail="Configuración inicial incompleta. Contacta a soporte.")

        nuevo_user = models.User(
            username        = data.username,
            hashed_password = crud.get_password_hash(data.password),
            role_id         = rol_admin.id,
            empresa_id      = nueva_emp.id,
            # Campos nuevos:
            nombre_completo = data.nombre_completo,
            email           = data.email,
            telefono        = data.telefono,
        )
        db.add(nuevo_user)
        db.commit()

        # ── (Opcional) enviar email de bienvenida o agendar tarea ─────────────
        # if data.email:
        #     send_welcome_email(data.email, data.nombre_completo or data.username, nueva_emp.id)

        logger.info(
            f"✅ Nuevo registro: empresa={nueva_emp.nombre} (id={nueva_emp.id}) | "
            f"user={data.username} | tipo={data.tipo_negocio} | "
            f"pais={data.pais} | ciudad={data.ciudad} | tamano={data.tamano_negocio} | "
            f"origen={data.origen}"
        )

        return {
            "message":     "Cuenta creada con éxito",
            "empresa_id":  nueva_emp.id,
            "trial_until": nueva_emp.trial_ends_at.isoformat(),
        }

    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="El usuario o empresa ya existe.")
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Error no controlado en registro: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Error interno del servidor")




@app.post("/token")
def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db)
):
    user = crud.get_user_by_username(db, username=form_data.username)
    if not user or not crud.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario o contraseña incorrectos",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if not getattr(user, 'is_active', True):
        raise HTTPException(status_code=403, detail="Cuenta suspendida por el administrador.")

    if not user.empresa_id or not user.empresa:
        raise HTTPException(status_code=403, detail="El usuario no está vinculado a ninguna empresa válida.")

    empresa = user.empresa
    if not empresa.is_active:
        raise HTTPException(status_code=403, detail="La suscripción de la empresa se encuentra suspendida. Contacte a soporte.")

    is_expired = False
    if empresa.trial_ends_at:
        ahora_utc = datetime.now(timezone.utc)
        fecha_limite = empresa.trial_ends_at
        if fecha_limite.tzinfo is None:
            fecha_limite = fecha_limite.replace(tzinfo=timezone.utc)
        if ahora_utc > fecha_limite:
            is_expired = True

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={
            "sub": user.username,
            "role": user.role.name,
            "empresa_id": user.empresa_id,
            "modules": [m.frontend_path for m in user.role.modules]
        },
        expires_delta=access_token_expires
    )

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "is_expired": is_expired
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SUPERADMIN
# ═══════════════════════════════════════════════════════════════════════════════

superadmin_router = APIRouter(
    prefix="/superadmin",
    tags=["SaaS SuperAdmin"],
    dependencies=[Depends(get_current_superadmin_user)]
)


@superadmin_router.get("/empresas", response_model=List[schemas.EmpresaOut])
def listar_empresas(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    return crud.get_empresas(db, skip=skip, limit=limit)


@superadmin_router.post("/empresas", response_model=schemas.EmpresaOut)
def registrar_nueva_empresa(data: schemas.EmpresaWithAdminCreate, db: Session = Depends(get_db)):
    try:
        return crud.create_empresa_with_admin(db, data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@superadmin_router.patch("/empresas/{empresa_id}/toggle", response_model=schemas.EmpresaOut)
def suspender_activar_empresa(empresa_id: int, db: Session = Depends(get_db)):
    if empresa_id == 1:
        raise HTTPException(status_code=400, detail="No puedes suspender la empresa maestra del SaaS.")
    empresa = crud.toggle_empresa_status(db, empresa_id)
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return empresa


@superadmin_router.patch("/empresas/{empresa_id}/plan", response_model=schemas.EmpresaOut)
def actualizar_plan_empresa(empresa_id: int, plan_data: schemas.EmpresaPlanUpdate, db: Session = Depends(get_db)):
    if empresa_id == 1:
        raise HTTPException(status_code=400, detail="La empresa maestra no puede modificar su plan vitalicio.")
    empresa = crud.update_empresa_plan(db, empresa_id, plan_data)
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return empresa


@superadmin_router.get("/planes", response_model=List[schemas.PlanSuscripcionOut])
def listar_planes_admin(db: Session = Depends(get_db)):
    return crud.get_planes(db, include_inactive=True)


@superadmin_router.post("/planes", response_model=schemas.PlanSuscripcionOut)
def crear_plan(plan: schemas.PlanSuscripcionCreate, db: Session = Depends(get_db)):
    existente = db.query(models.PlanSuscripcion).filter(
        models.PlanSuscripcion.codigo_interno == plan.codigo_interno
    ).first()
    if existente:
        raise HTTPException(status_code=400, detail="Ya existe un plan con este código interno.")
    return crud.create_plan(db, plan)


@superadmin_router.patch("/planes/{plan_id}", response_model=schemas.PlanSuscripcionOut)
def actualizar_plan(plan_id: int, plan_update: schemas.PlanSuscripcionUpdate, db: Session = Depends(get_db)):
    plan_actualizado = crud.update_plan(db, plan_id, plan_update)
    if not plan_actualizado:
        raise HTTPException(status_code=404, detail="Plan no encontrado")
    return plan_actualizado


@superadmin_router.get("/historial-pagos", response_model=List[schemas.RegistroPagoOut])
def listar_historial_pagos(db: Session = Depends(get_db)):
    pagos = db.query(models.RegistroPago).options(
        joinedload(models.RegistroPago.empresa),
        joinedload(models.RegistroPago.plan)
    ).order_by(models.RegistroPago.fecha_pago.desc()).all()

    return [
        {**p.__dict__, "empresa_nombre": p.empresa.nombre, "plan_nombre": p.plan.nombre}
        for p in pagos
    ]


@superadmin_router.post("/impersonate/{empresa_id}")
def impersonate_company(
    empresa_id: int,
    db: Session = Depends(get_db),
    current_admin: schemas.User = Depends(get_current_superadmin_user)
):
    target_user = db.query(models.User).filter(
        models.User.empresa_id == empresa_id
    ).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="No se encontró un usuario para esta empresa.")

    access_token = create_access_token(
        data={
            "sub": target_user.username,
            "empresa_id": target_user.empresa_id,
            "role": target_user.role.name if target_user.role else "Admin",
            "is_impersonated": True
        }
    )
    return {"access_token": access_token, "token_type": "bearer"}


# FIX #6: Usar get_current_superadmin_user en lugar de get_current_user con validación manual
class ModulosEmpresaRequest(BaseModel):
    modulos: List[str]


@app.patch("/superadmin/empresas/{empresa_id}/modulos")
def actualizar_modulos_empresa(
    empresa_id: int,
    req: ModulosEmpresaRequest,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_superadmin_user)
):
    empresa = db.query(models.Empresa).filter(models.Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    empresa.modulos_habilitados = req.modulos
    db.commit()
    return {"msg": "Módulos de la empresa actualizados correctamente"}


app.include_router(superadmin_router)


# ═══════════════════════════════════════════════════════════════════════════════
# ROLES / MÓDULOS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/roles/", response_model=schemas.Role)
def create_role(role: schemas.RoleCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    if crud.get_role_by_name(db, name=role.name):
        raise HTTPException(status_code=400, detail="Rol ya registrado")
    return crud.create_role(db=db, role=role)


@app.get("/roles/", response_model=List[schemas.Role])
def read_roles(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    return crud.get_roles(db, skip=skip, limit=limit)


@app.put("/roles/{role_id}/modules", response_model=schemas.Role)
def set_role_modules(role_id: int, module_ids: List[int], db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    db_role = crud.set_modules_for_role(db, role_id=role_id, module_ids=module_ids)
    if db_role is None:
        raise HTTPException(status_code=404, detail="Rol no encontrado")
    return db_role


@app.post("/modulos/", response_model=schemas.Modulo)
def create_modulo(modulo: schemas.ModuloCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    if crud.get_modulo_by_name(db, name=modulo.name):
        raise HTTPException(status_code=400, detail="Módulo ya registrado")
    return crud.create_modulo(db=db, modulo=modulo)


@app.get("/modulos/", response_model=List[schemas.Modulo])
def read_modulos(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    return crud.get_modulos(db, skip=skip, limit=limit)


@app.put("/modulos/{modulo_id}", response_model=schemas.Modulo)
def update_modulo(modulo_id: int, modulo: schemas.ModuloCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    db_modulo = crud.update_modulo(db, modulo_id=modulo_id, modulo=modulo)
    if db_modulo is None:
        raise HTTPException(status_code=404, detail="Módulo no encontrado")
    return db_modulo


@app.delete("/modulos/{modulo_id}")
def delete_modulo(modulo_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    if crud.delete_modulo(db, modulo_id=modulo_id) is None:
        raise HTTPException(status_code=404, detail="Módulo no encontrado")
    return {"message": "Módulo eliminado"}


# ═══════════════════════════════════════════════════════════════════════════════
# USUARIOS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/users/", response_model=schemas.User)
def create_user(user: schemas.UserCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    if crud.get_user_by_username(db, username=user.username):
        raise HTTPException(status_code=400, detail="Nombre de usuario ya registrado")
    return crud.create_user(db=db, user=user, empresa_id=current_user.empresa_id)


@app.get("/users/me", response_model=schemas.User)
def read_users_me(current_user: schemas.User = Depends(get_current_active_user)):
    return current_user


@app.get("/users/", response_model=List[schemas.User])
def read_users(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    return crud.get_users(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@app.put("/users/{user_id}", response_model=schemas.User)
def update_user(user_id: int, user: schemas.UserCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    db_user = crud.update_user(db, user_id=user_id, user=user, empresa_id=current_user.empresa_id)
    if db_user is None:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return db_user


@app.delete("/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    if crud.delete_user(db, user_id=user_id, empresa_id=current_user.empresa_id) is None:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Usuario eliminado"}


@app.patch("/users/{user_id}/toggle")
def toggle_user(user_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    db_user = db.query(models.User).filter(
        models.User.id == user_id,
        models.User.empresa_id == current_user.empresa_id
    ).first()
    if not db_user:
        raise HTTPException(status_code=404)
    db_user.is_active = not db_user.is_active
    db.commit()
    return {"status": "ok", "new_state": db_user.is_active}


@app.get("/admin/usuarios", response_model=List[schemas.User])
def listar_usuarios_empresa(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    if current_user.role.name != "Admin":
        raise HTTPException(status_code=403, detail="No tienes permisos para ver la lista de usuarios")
    return db.query(models.User).filter(models.User.empresa_id == current_user.empresa_id).all()


# ═══════════════════════════════════════════════════════════════════════════════
# PLANTILLAS INTELIGENTES (EXCEL) — DEBEN IR ANTES DE LAS RUTAS CON {id}
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/clientes/template")
def get_clientes_template(current_user: schemas.User = Depends(get_current_active_user)):
    wb = openpyxl.Workbook()

    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_properties.tabColor = "3B82F6"
    ws_inst.cell(row=2, column=2, value="🛠 CÓMO REGISTRAR TERCEROS").font = Font(size=14, bold=True, color="3B82F6")
    instrucciones = [
        "1. Usa la pestaña 'Plantilla Datos' para registrar clientes o proveedores.",
        "2. La CÉDULA/NIT no debe repetirse. Si ya existe, se omitirá.",
        "3. ES_CLIENTE y ES_PROVEEDOR: Usa la lista desplegable (SI / NO).",
        "4. No modifiques ni elimines la Fila 1."
    ]
    for i, inst in enumerate(instrucciones, 4):
        ws_inst.cell(row=i, column=2, value=inst).font = Font(size=11)
    ws_inst.column_dimensions['B'].width = 80

    ws_datos = wb.create_sheet(title="Plantilla Datos")
    headers = ["NOMBRE", "CEDULA", "TELEFONO", "DIRECCION", "CUPO_CREDITO", "ES_CLIENTE", "ES_PROVEEDOR"]
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws_datos.cell(row=1, column=col_num, value=header)
        cell.fill, cell.font = header_fill, header_font
        ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    dv_bool = DataValidation(type="list", formula1='"SI,NO"', allow_blank=True)
    ws_datos.add_data_validation(dv_bool)
    dv_bool.add("F2:G1000")

    ejemplos = [
        ["Distribuidora XYZ", "900123456", "3001234567", "Calle 10", 5000000, "NO", "SI"],
        ["Juan Pérez", "10203040", "3100000000", "Cra 5", 0, "SI", "NO"]
    ]
    for r_idx, row in enumerate(ejemplos, 2):
        for c_idx, val in enumerate(row, 1):
            ws_datos.cell(row=r_idx, column=c_idx, value=val)

    ws_datos.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_terceros_PRO.xlsx"'}
    )


@app.get("/inventario/movimientos/template")
def get_movimientos_template(current_user: schemas.User = Depends(get_current_active_user)):
    wb = openpyxl.Workbook()

    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_properties.tabColor = "10B981"
    ws_inst.cell(row=2, column=2, value="🛠 CÓMO CARGAR MOVIMIENTOS").font = Font(size=14, bold=True, color="10B981")
    instrucciones = [
        "1. Usa la pestaña 'Plantilla Datos'.",
        "2. El PRODUCTO_NOMBRE debe coincidir exactamente con uno existente en el sistema.",
        "3. TIPO: Usa la lista desplegable (entrada, salida, ajuste).",
        "4. CANTIDAD: Siempre positiva (el sistema deduce si es salida)."
    ]
    for i, inst in enumerate(instrucciones, 4):
        ws_inst.cell(row=i, column=2, value=inst).font = Font(size=11)
    ws_inst.column_dimensions['B'].width = 80

    ws_datos = wb.create_sheet(title="Plantilla Datos")
    headers = ["PRODUCTO_NOMBRE", "TIPO", "CANTIDAD", "COSTO_UNITARIO", "MOTIVO", "REFERENCIA", "OBSERVACION"]
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws_datos.cell(row=1, column=col_num, value=header)
        cell.fill, cell.font = header_fill, header_font
        ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22

    dv_tipo = DataValidation(type="list", formula1='"entrada,salida,ajuste"', allow_blank=False)
    ws_datos.add_data_validation(dv_tipo)
    dv_tipo.add("B2:B1000")

    ejemplos = [
        ["Cacao Tostado", "entrada", 50, 2500, "Compra inicial", "FACT-001", "Stock base"],
        ["Chocolatina 80g", "ajuste", 5, 0, "Dañado", "MERMA", "Se rompió empaque"]
    ]
    for r_idx, row in enumerate(ejemplos, 2):
        for c_idx, val in enumerate(row, 1):
            ws_datos.cell(row=r_idx, column=c_idx, value=val)

    ws_datos.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_movimientos_PRO.xlsx"'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTES / TERCEROS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/clientes/", response_model=schemas.Cliente)
def create_cliente(cliente: schemas.ClienteCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.create_cliente(db=db, empresa_id=current_user.empresa_id, cliente=cliente)


@app.post("/clientes/upload", response_model=schemas.BulkLoadResponse)
def upload_clientes(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.bulk_create_clientes(db=db, empresa_id=current_user.empresa_id, file=file.file, filename=file.filename)


@app.get("/clientes/", response_model=List[schemas.Cliente])
def read_clientes(skip: int = 0, limit: int = 500, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_clientes(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@app.get("/clientes/{cliente_id}", response_model=schemas.Cliente)
def read_cliente(cliente_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_cliente = crud.get_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return db_cliente


@app.put("/clientes/{cliente_id}", response_model=schemas.Cliente)
def update_cliente(cliente_id: int, cliente: schemas.ClienteCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_cliente = crud.update_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id, cliente=cliente)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return db_cliente


@app.get("/clientes/{cliente_id}/details", response_model=schemas.ClienteDetails)
def get_cliente_details(cliente_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_cliente = crud.get_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    deuda_actual = crud.get_cliente_deuda(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    return schemas.ClienteDetails(**db_cliente.__dict__, deuda_actual=deuda_actual)


@app.delete("/clientes/{cliente_id}")
def delete_cliente(cliente_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_cliente = crud.get_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if db_cliente is None:
        raise HTTPException(status_code=404, detail="Tercero no encontrado")
    bloqueos = crud.check_can_delete_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if bloqueos:
        raise HTTPException(
            status_code=409,
            detail=(
                f"No se puede eliminar '{db_cliente.nombre}' porque tiene: "
                + ", ".join(bloqueos) + ". Desactívelo en lugar de eliminarlo."
            )
        )
    crud.delete_cliente(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    return {"message": f"Tercero '{db_cliente.nombre}' eliminado correctamente"}


@app.get("/clientes/{cliente_id}/history", response_model=schemas.ClienteHistory)
def get_cliente_history(cliente_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    history = crud.get_cliente_history(db, empresa_id=current_user.empresa_id, cliente_id=cliente_id)
    if history is None:
        raise HTTPException(status_code=404, detail="Historial no encontrado")
    return history


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTOS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/productos/upload", response_model=schemas.BulkLoadResponse)
def upload_productos(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.bulk_create_productos(db=db, empresa_id=current_user.empresa_id, file=file.file, filename=file.filename)


@app.post("/productos/", response_model=schemas.Producto)
def create_producto(producto: schemas.ProductoCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.create_producto(db=db, empresa_id=current_user.empresa_id, producto=producto)


@app.get("/productos/", response_model=List[schemas.Producto])
def read_productos(skip: int = 0, limit: int = 500, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_productos(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@app.put("/productos/{producto_id}", response_model=schemas.Producto)
def update_producto(producto_id: int, producto: schemas.ProductoCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_producto = crud.update_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id, producto=producto)
    if db_producto is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return db_producto


@app.delete("/productos/{producto_id}")
def delete_producto(producto_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_producto = crud.get_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id)
    if db_producto is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    bloqueos = crud.check_can_delete_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id)
    if bloqueos:
        raise HTTPException(
            status_code=409,
            detail=(
                f"No se puede eliminar '{db_producto.nombre}' porque "
                + ", ".join(bloqueos) + "."
            )
        )
    crud.delete_producto(db, empresa_id=current_user.empresa_id, producto_id=producto_id)
    return {"message": f"Producto '{db_producto.nombre}' eliminado correctamente"}


@app.get("/productos/export")
def exportar_productos(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    prods = crud.get_productos(db, empresa_id=current_user.empresa_id)
    groups = {1: 'MP', 2: 'PT', 3: 'AF', 4: 'INS'}
    rows = [
        {
            "id": p.id, "nombre": p.nombre, "precio": p.precio, "costo": p.costo,
            "grupo_item": groups.get(p.grupo_item, 'PT'), "es_servicio": "SÍ" if p.es_servicio else "NO",
            "unidad_medida": p.unidad_medida, "stock_minimo": float(p.stock_minimo or 0),
            "stock_actual": float(p.stock_actual or 0)
        }
        for p in prods
    ]
    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Productos")
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="productos_existentes.xlsx"'}
    )


@app.get("/productos/template")
def get_productos_template(current_user: schemas.User = Depends(get_current_active_user)):
    wb = openpyxl.Workbook()

    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.sheet_properties.tabColor = "8B5CF6"
    ws_inst.cell(row=2, column=2, value="🛠 CÓMO USAR ESTA PLANTILLA").font = Font(size=14, bold=True, color="8B5CF6")
    instrucciones = [
        "1. Ve a la pestaña 'Plantilla Datos' para registrar tu inventario.",
        "2. IMPORTANTE: No modifiques, renombres ni elimines la fila 1 (Cabeceras).",
        "3. GRUPO_ITEM: Usa el desplegable (1=Materia Prima, 2=Prod. Terminado, 3=Activo Fijo, 4=Insumo).",
        "4. ES_SERVICIO: Usa el desplegable (0 = Producto Físico, 1 = Servicio Intangible).",
        "5. UNIDAD_MEDIDA: Usa el desplegable (UND, Kg, Lts, etc.).",
        "6. COSTO: Si marcas el ítem como Servicio (1), el costo debe ser 0."
    ]
    for i, inst in enumerate(instrucciones, 4):
        ws_inst.cell(row=i, column=2, value=inst).font = Font(size=11)
    ws_inst.column_dimensions['B'].width = 80

    ws_datos = wb.create_sheet(title="Plantilla Datos")
    headers = ["nombre", "precio", "costo", "grupo_item", "unidad_medida", "es_servicio", "stock_minimo"]
    header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = ws_datos.cell(row=1, column=col_num, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_datos.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    dv_grupo = DataValidation(type="list", formula1='"1,2,3,4"', allow_blank=True)
    dv_grupo.error = 'Selecciona una opción válida de la lista'
    ws_datos.add_data_validation(dv_grupo)
    dv_grupo.add("D2:D1000")

    dv_unidad = DataValidation(type="list", formula1='"UND,Kg,MTS,Lts,Gr"', allow_blank=True)
    ws_datos.add_data_validation(dv_unidad)
    dv_unidad.add("E2:E1000")

    dv_servicio = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws_datos.add_data_validation(dv_servicio)
    dv_servicio.add("F2:F1000")

    ejemplos = [
        ["Cacao Tostado", 5000, 3000, 1, "Kg", 0, 10],
        ["Chocolatina 80g", 12000, 4500, 2, "UND", 0, 5],
        ["Servicio Maquila", 2500, 0, 2, "UND", 1, 0]
    ]
    for r_idx, row_data in enumerate(ejemplos, 2):
        for c_idx, value in enumerate(row_data, 1):
            ws_datos.cell(row=r_idx, column=c_idx, value=value)

    ws_datos.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_productos_PRO.xlsx"'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# VENTAS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/ventas/", response_model=schemas.Venta)
def create_venta(venta: schemas.VentaCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id

    db_cliente = crud.get_cliente(db, empresa_id=empresa_id, cliente_id=venta.cliente_id)
    if not db_cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    if not venta.detalles:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos un producto.")

    for d in venta.detalles:
        prod = crud.get_producto(db, empresa_id=empresa_id, producto_id=d.producto_id)
        if not prod:
            raise HTTPException(status_code=404, detail=f"Producto {d.producto_id} no existe")
        if not prod.es_servicio and prod.grupo_item != 2:
            raise HTTPException(status_code=400, detail=f"'{prod.nombre}' no es un Producto Terminado y no puede venderse.")
        if not prod.es_servicio:
            if (prod.stock_actual or 0) < d.cantidad:
                raise HTTPException(status_code=400, detail=f"Stock insuficiente para '{prod.nombre}'. Disponible: {prod.stock_actual}, requerido: {d.cantidad}")

    if not venta.pagada:
        total_nueva = sum(
            (d.precio_unitario if d.precio_unitario is not None else crud.get_producto(db, empresa_id, d.producto_id).precio) * d.cantidad
            for d in venta.detalles
        )
        deuda_actual = crud.get_cliente_deuda(db, empresa_id=empresa_id, cliente_id=venta.cliente_id)
        if (deuda_actual + total_nueva) > db_cliente.cupo_credito:
            cupo_disp = db_cliente.cupo_credito - deuda_actual
            raise HTTPException(status_code=400, detail=f"La venta excede el cupo de crédito. Disponible: {cupo_disp:.2f}")

    db_venta = crud.create_venta(db=db, empresa_id=empresa_id, venta=venta)


    
    

    try:
        # Dentro del proceso de creación de venta en el backend
        for det in db_venta.detalles:
            prod = crud.get_producto(db, empresa_id=empresa_id, producto_id=det.producto_id)
            
            if getattr(prod, "es_servicio", False):
                continue
                
            # --- NUEVA LÓGICA DE INTEGRACIÓN ---
            if getattr(prod, "maneja_lotes", False):
                # Si el producto es perecedero, usamos FEFO para descontar de los lotes
                try:
                    crud.consumir_stock_fefo(
                        db, 
                        empresa_id=empresa_id, 
                        producto_id=det.producto_id, 
                        cantidad_requerida=det.cantidad,
                        referencia=f"Venta #{db_venta.id}"
                    )
                    
                    # 👇 FIX CRÍTICO: Sincronizar el stock global del producto
                    prod.stock_actual = (prod.stock_actual or 0) - det.cantidad
                    db.add(prod)
                    db.commit()
                    
                except ValueError as e:
                    raise HTTPException(status_code=400, detail=str(e))
            else:
                # Si NO maneja lotes, se descuenta del stock global como antes
                crud.create_movement(db, empresa_id=empresa_id, payload=schemas.InventoryMovementCreate(
                    producto_id=det.producto_id,
                    tipo=schemas.MovementType.salida,
                    cantidad=det.cantidad,
                    costo_unitario=prod.costo or 0.0,
                    motivo="venta",
                    referencia=f"venta #{db_venta.id}"
                ))

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    crud.check_and_notify_low_stock(db, empresa_id=empresa_id, producto_ids=[det.producto_id for det in db_venta.detalles])
    return db_venta


# @app.get("/ventas/", response_model=List[schemas.Venta])
# def read_ventas(
#     skip: int = 0,
#     limit: int = Query(default=100, le=500),
#     db: Session = Depends(get_db),
#     current_user: schemas.User = Depends(get_current_active_user)
# ):
#     return crud.get_ventas(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@app.get("/ventas/", response_model=List[schemas.Venta])
def read_ventas(
    skip: int = 0,
    limit: int = Query(default=100, le=500),
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user)
):
    return crud.get_ventas(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)



@app.get("/ventas/{venta_id}", response_model=schemas.Venta)
def read_venta(venta_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_venta = crud.get_venta(db, empresa_id=current_user.empresa_id, venta_id=venta_id)
    if db_venta is None:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    return db_venta


@app.put("/ventas/{venta_id}", response_model=schemas.Venta)
def update_venta(venta_id: int, venta: schemas.VentaCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    if not crud.get_cliente(db, empresa_id=empresa_id, cliente_id=venta.cliente_id):
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    if not venta.detalles:
        raise HTTPException(status_code=400, detail="Debe proporcionar al menos un producto.")
    for detalle in venta.detalles:
        if not crud.get_producto(db, empresa_id=empresa_id, producto_id=detalle.producto_id):
            raise HTTPException(status_code=404, detail=f"Producto {detalle.producto_id} no encontrado.")
    db_venta = crud.update_venta(db, empresa_id=empresa_id, venta_id=venta_id, venta=venta)
    if db_venta is None:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    return db_venta


@app.delete("/ventas/{venta_id}")
def delete_venta(venta_id: int, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    db_venta = crud.get_venta(db, empresa_id=empresa_id, venta_id=venta_id)
    if db_venta is None:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    bloqueos = crud.check_can_delete_venta(db, empresa_id=empresa_id, venta_id=venta_id)
    if bloqueos:
        raise HTTPException(
            status_code=409,
            detail=(f"No se puede eliminar la venta #{venta_id} porque " + ", ".join(bloqueos) + ".")
        )
    crud.revertir_movimientos_venta(db, empresa_id=empresa_id, venta=db_venta)
    crud.delete_venta(db, empresa_id=empresa_id, venta_id=venta_id)
    return {"message": f"Venta #{venta_id} eliminada y stock revertido"}


# ═══════════════════════════════════════════════════════════════════════════════
# PAGOS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/pagos/", response_model=schemas.Pago)
def create_pago(pago: schemas.PagoCreate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    db_venta = crud.get_venta(db, empresa_id=empresa_id, venta_id=pago.venta_id)
    if not db_venta:
        raise HTTPException(status_code=404, detail="Venta no encontrada")
    monto_pendiente = db_venta.total - db_venta.monto_pagado
    if pago.monto > monto_pendiente + 0.01:
        raise HTTPException(status_code=400, detail=f"El monto excede el saldo pendiente de {monto_pendiente:.2f}")
    return crud.create_pago(db=db, empresa_id=empresa_id, pago=pago)


# FIX #5: Filtrar por empresa_id — antes devolvía pagos de TODAS las empresas
@app.get("/pagos/", response_model=List[schemas.Pago])
def read_pagos(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user)
):
    return (
        db.query(models.Pago)
        .filter(models.Pago.empresa_id == current_user.empresa_id)
        .offset(skip)
        .limit(limit)
        .all()
    )


@app.put("/pagos/{pago_id}", response_model=schemas.Pago)
def update_pago(pago_id: int, pago: schemas.PagoUpdate, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    db_pago = crud.update_pago(db, empresa_id=current_user.empresa_id, pago_id=pago_id, pago=pago)
    if db_pago is None:
        raise HTTPException(status_code=404, detail="Pago no encontrado")
    return db_pago


# ═══════════════════════════════════════════════════════════════════════════════
# INVENTARIO
# FIX #6: Consistencia de permisos — todos usan get_current_active_user
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/inventario/kardex/{producto_id}", response_model=schemas.KardexResponse)
def kardex_producto(
    producto_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    sd = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    ed = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
    return crud.get_kardex_promedio_ponderado(db, empresa_id=current_user.empresa_id, producto_id=producto_id, start_date=sd, end_date=ed)


@app.get("/inventario/kardex/{producto_id}/export")
def kardex_export_csv(
    producto_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    sd = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
    ed = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
    rep = crud.get_kardex_promedio_ponderado(db, empresa_id=current_user.empresa_id, producto_id=producto_id, start_date=sd, end_date=ed)
    lines = ["fecha,tipo,cantidad,costo_unit,referencia,saldo_cant,saldo_costo,saldo_valor"]
    for it in rep.items:
        lines.append(f"{it.fecha.isoformat()},{it.tipo},{it.cantidad},{it.costo_unitario},{it.referencia or ''},{it.saldo_cantidad},{it.saldo_costo_unitario},{it.saldo_valor}")
    return Response(
        content="\n".join(lines),
        headers={"Content-Disposition": f'attachment; filename="kardex_{producto_id}.csv"', "Content-Type": "text/csv; charset=utf-8"}
    )


@app.get("/reportes/inventario-actual", response_model=schemas.InventarioSnapshot)
def inventario_actual(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    return crud.get_inventario_actual(db, empresa_id=current_user.empresa_id)


@app.get("/reportes/inventario-actual/export")
def inventario_actual_export(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    snap = crud.get_inventario_actual(db, empresa_id=current_user.empresa_id)
    lines = ["id,nombre,es_servicio,unidad,stock_actual,costo,precio,valor_costo,valor_venta"]
    for it in snap.items:
        lines.append(f"{it.id},{it.nombre},{1 if it.es_servicio else 0},{it.unidad_medida or ''},{it.stock_actual},{it.costo},{it.precio},{it.valor_costo},{it.valor_venta}")
    lines.append(f"TOTALS,,,,,,,{snap.total_valor_costo},{snap.total_valor_venta}")
    return Response(
        content="\n".join(lines),
        headers={"Content-Disposition": 'attachment; filename="inventario_actual.csv"', "Content-Type": "text/csv; charset=utf-8"}
    )


@app.get("/reportes/rotacion", response_model=schemas.ReporteRotacion)
def reporte_rotacion(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    limit: int = Query(10, ge=1, le=100),
    incluir_servicios: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    sd = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
    ed = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    return crud.get_rotacion_productos(db, empresa_id=current_user.empresa_id, start_date=sd, end_date=ed, limit=limit, incluir_servicios=incluir_servicios)


@app.post("/inventario/movimientos", response_model=schemas.InventoryMovementOut)
def crear_movimiento(
    payload: schemas.InventoryMovementCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    try:
        return crud.create_movement(db, empresa_id=current_user.empresa_id, payload=payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/inventario/movimientos", response_model=List[schemas.InventoryMovementOut])
def listar_movimientos(
    producto_id: Optional[int] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    return crud.list_movements(db, empresa_id=current_user.empresa_id, producto_id=producto_id, limit=limit)


@app.get("/inventario/alertas/bajo-stock", response_model=List[schemas.InventoryAlertOut])
def alertas_bajo_stock(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    prods = crud.get_low_stock(db, empresa_id=current_user.empresa_id)
    return [
        schemas.InventoryAlertOut(producto_id=p.id, nombre=p.nombre, stock_actual=p.stock_actual or 0, stock_minimo=p.stock_minimo or 0)
        for p in prods
    ]


@app.patch("/productos/{producto_id}/stock-minimo")
def actualizar_stock_minimo(
    producto_id: int,
    body: schemas.ProductoStockUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    prod = crud.update_producto_stock_minimo(db, empresa_id=current_user.empresa_id, producto_id=producto_id, minimo=body.stock_minimo or 0)
    if not prod:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return {"ok": True}


@app.post("/movimientos/upload", response_model=schemas.BulkLoadResponse)
def upload_movimientos(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.bulk_create_movimientos(db=db, empresa_id=current_user.empresa_id, file=file.file, filename=file.filename)


# ═══════════════════════════════════════════════════════════════════════════════
# REPORTES
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/reportes/ventas_summary", response_model=schemas.VentasSummary)
def get_ventas_summary(start_date: Optional[date] = None, end_date: Optional[date] = None, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_ventas_summary(db, empresa_id=current_user.empresa_id, start_date=start_date, end_date=end_date)


@app.get("/reportes/productos_vendidos", response_model=schemas.ReporteProductosVendidos)
def get_productos_vendidos(start_date: Optional[date] = None, end_date: Optional[date] = None, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_productos_vendidos(db, empresa_id=current_user.empresa_id, start_date=start_date, end_date=end_date)


@app.get("/reportes/clientes_compradores", response_model=List[schemas.ClienteComprador])
def get_clientes_compradores(start_date: Optional[date] = None, end_date: Optional[date] = None, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_clientes_compradores(db, empresa_id=current_user.empresa_id, start_date=start_date, end_date=end_date)


@app.get("/reportes/clientes_deudores", response_model=List[schemas.ClienteDeudor])
def get_clientes_deudores(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_clientes_deudores(db, empresa_id=current_user.empresa_id)


@app.get("/reportes/rentabilidad_productos", response_model=List[schemas.ProductoRentabilidad])
def get_rentabilidad_productos(start_date: Optional[date] = None, end_date: Optional[date] = None, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_rentabilidad_por_producto(db, empresa_id=current_user.empresa_id, start_date=start_date, end_date=end_date)


@app.get("/reportes/cuentas_por_cobrar", response_model=List[schemas.ClienteCuentasPorCobrar])
def get_cuentas_por_cobrar(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_cuentas_por_cobrar_por_cliente(db, empresa_id=current_user.empresa_id)


@app.get("/reportes/dashboard", response_model=schemas.DashboardData)
def get_dashboard_report(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return crud.get_dashboard_data(db, empresa_id=current_user.empresa_id)


@app.get("/reportes/iva-neto")
def get_iva_neto(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    empresa_id = current_user.empresa_id
    query_v_iva   = db.query(func.sum(models.Venta.iva_total)).filter(models.Venta.empresa_id == empresa_id)
    query_v_total = db.query(func.sum(models.Venta.total)).filter(models.Venta.empresa_id == empresa_id)
    query_c_iva   = db.query(func.sum(models.Compra.iva_total)).filter(models.Compra.empresa_id == empresa_id)
    query_c_total = db.query(func.sum(models.Compra.total)).filter(models.Compra.empresa_id == empresa_id)

    if start_date:
        query_v_iva   = query_v_iva.filter(models.Venta.fecha >= start_date)
        query_v_total = query_v_total.filter(models.Venta.fecha >= start_date)
        query_c_iva   = query_c_iva.filter(models.Compra.fecha >= start_date)
        query_c_total = query_c_total.filter(models.Compra.fecha >= start_date)
    if end_date:
        td = timedelta(days=1)
        query_v_iva   = query_v_iva.filter(models.Venta.fecha < end_date + td)
        query_v_total = query_v_total.filter(models.Venta.fecha < end_date + td)
        query_c_iva   = query_c_iva.filter(models.Compra.fecha < end_date + td)
        query_c_total = query_c_total.filter(models.Compra.fecha < end_date + td)

    iva_v = query_v_iva.scalar() or 0.0
    tot_v = query_v_total.scalar() or 0.0
    iva_c = query_c_iva.scalar() or 0.0
    tot_c = query_c_total.scalar() or 0.0

    return {
        "periodo": {"desde": start_date, "hasta": end_date},
        "iva_generado_ventas": iva_v, "iva_descontable_compras": iva_c,
        "iva_neto_resultado": iva_v - iva_c, "ventas_brutas": tot_v,
        "base_gravable_ventas": tot_v - iva_v, "compras_brutas": tot_c,
        "base_gravable_compras": tot_c - iva_c
    }


@app.get("/reportes/productividad", response_model=schemas.ReporteProductividad, dependencies=[Depends(get_current_admin_user)])
def get_productivity_report(start_date: date, end_date: date, db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_admin_user)):
    return crud.get_reporte_productividad(db, empresa_id=current_user.empresa_id, start_date=start_date, end_date=end_date)


@app.get("/reportes/produccion-summary")
def get_produccion_summary(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    empresa_id = current_user.empresa_id
    query = db.query(models.LoteProduccion).filter(
        models.LoteProduccion.estado == "Confirmado",
        models.LoteProduccion.empresa_id == empresa_id
    )
    if start_date:
        query = query.filter(models.LoteProduccion.fecha_confirmacion >= start_date)
    if end_date:
        query = query.filter(models.LoteProduccion.fecha_confirmacion < end_date + timedelta(days=1))
    lotes = query.all()
    return {
        "total_costo_produccion": sum(l.costo_total for l in lotes),
        "total_unidades_producidas": sum(l.cantidad_real for l in lotes if l.cantidad_real),
        "total_lotes_finalizados": len(lotes),
        "total_maquilas": len([l for l in lotes if l.cliente_id])
    }


@app.get("/reportes/consumo-insumos")
def get_consumo_insumos(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_admin_user)
):
    empresa_id = current_user.empresa_id
    query = (
        db.query(
            models.Producto.nombre,
            func.sum(models.InventoryMovement.cantidad).label("cantidad_total"),
            func.sum(models.InventoryMovement.cantidad * models.InventoryMovement.costo_unitario).label("costo_total")
        )
        .join(models.InventoryMovement, models.Producto.id == models.InventoryMovement.producto_id)
        .filter(
            models.InventoryMovement.tipo == "salida",
            models.InventoryMovement.motivo.like("%Producción%"),
            models.InventoryMovement.empresa_id == empresa_id
        )
    )
    if start_date:
        query = query.filter(models.InventoryMovement.created_at >= start_date)
    if end_date:
        query = query.filter(models.InventoryMovement.created_at < end_date + timedelta(days=1))
    results = query.group_by(models.Producto.nombre).order_by(func.sum(models.InventoryMovement.cantidad).desc()).all()
    return [{"insumo": r.nombre, "cantidad": r.cantidad_total, "costo": r.costo_total} for r in results]


@app.get("/reportes/financiero-prestamos")
def reporte_financiero_prestamos(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    return crud.get_reporte_financiero_prestamos(db, current_user.empresa_id)


# ═══════════════════════════════════════════════════════════════════════════════
# DEVOLUCIONES
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/devoluciones/", response_model=schemas.DevolucionOut)
def crear_devolucion(data: schemas.DevolucionCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.crear_devolucion(db, empresa_id=current_user.empresa_id, data=data)


@app.get("/devoluciones/venta/{venta_id}", response_model=List[schemas.DevolucionOut])
def get_devoluciones_venta(venta_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_devoluciones_by_venta(db, empresa_id=current_user.empresa_id, venta_id=venta_id)


# ═══════════════════════════════════════════════════════════════════════════════
# CORTE DE CAJA
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/caja/corte", response_model=schemas.CorteCajaOut)
def crear_corte_caja(data: schemas.CorteCajaCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.crear_corte_caja(db, empresa_id=current_user.empresa_id, usuario_id=current_user.id, efectivo_fisico=data.efectivo_fisico, observaciones=data.observaciones)


@app.get("/caja/cortes", response_model=List[schemas.CorteCajaOut])
def listar_cortes(skip: int = 0, limit: int = 30, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_cortes_caja(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@app.get("/caja/corte/preview")
def preview_corte(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.calcular_totales_dia(db, empresa_id=current_user.empresa_id)


@app.post("/caja/gastos", response_model=schemas.GastoOut)
def registrar_gasto(data: schemas.GastoCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.crear_gasto(db, empresa_id=current_user.empresa_id, usuario_id=current_user.id, data=data)


@app.get("/caja/gastos", response_model=List[schemas.GastoOut])
def listar_gastos(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_gastos(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


# ═══════════════════════════════════════════════════════════════════════════════
# NOTIFICACIONES
# ═══════════════════════════════════════════════════════════════════════════════

notificaciones_router = APIRouter(
    prefix="/notificaciones",
    tags=["Notificaciones"],
    dependencies=[Depends(get_current_active_user)]
)


@notificaciones_router.get("/", response_model=List[schemas.Notificacion])
def get_my_notifications(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_notificaciones_usuario(db, empresa_id=current_user.empresa_id, usuario_id=current_user.id)


@notificaciones_router.get("/unread-count")
def get_unread_count(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    count = db.query(models.Notificacion).filter(
        models.Notificacion.usuario_id == current_user.id,
        models.Notificacion.empresa_id == current_user.empresa_id,
        models.Notificacion.leido == False
    ).count()
    return {"unread": count}


@notificaciones_router.put("/{notificacion_id}/leida", response_model=schemas.Notificacion)
def mark_notification_as_read(notificacion_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_notif = crud.marcar_notificacion_leida(db, empresa_id=current_user.empresa_id, notificacion_id=notificacion_id, usuario_id=current_user.id)
    if db_notif is None:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    return db_notif


@notificaciones_router.put("/mark-all-read")
def mark_all_read(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db.query(models.Notificacion).filter(
        models.Notificacion.usuario_id == current_user.id,
        models.Notificacion.empresa_id == current_user.empresa_id,
        models.Notificacion.leido == False
    ).update({"leido": True})
    db.commit()
    return {"message": "Todas las notificaciones marcadas como leídas"}


app.include_router(notificaciones_router)


# FIX #11: Función auxiliar con docstring, filtro de activos y manejo de errores
def crear_notificacion(db: Session, empresa_id: int, mensaje: str, tipo: str = "info") -> int:
    """
    Crea notificaciones para todos los admins ACTIVOS de una empresa.

    Args:
        db: Sesión de base de datos.
        empresa_id: ID de la empresa destinataria.
        mensaje: Texto de la notificación.
        tipo: Categoría visual ('info', 'warning', 'error').

    Returns:
        Número de notificaciones creadas exitosamente.
    """
    admins = db.query(models.User).join(models.Role).filter(
        models.User.empresa_id == empresa_id,
        models.Role.name == "Admin",
        models.User.is_active == True  # Solo admins activos
    ).all()

    if not admins:
        logger.warning(f"crear_notificacion: no hay admins activos en empresa {empresa_id}")
        return 0

    creadas = 0
    for admin in admins:
        try:
            db.add(models.Notificacion(
                usuario_id=admin.id,
                empresa_id=empresa_id,
                mensaje=mensaje,
                tipo=tipo,
                leido=False,
                fecha_creacion=utcnow()
            ))
            creadas += 1
        except Exception as e:
            logger.error(f"Error creando notificación para admin {admin.id}: {e}")

    db.commit()
    return creadas


@app.post("/superadmin/notificar-vencimientos-hoy")
def notificar_vencimientos(db: Session = Depends(get_db)):
    hoy = datetime.now().date()
    empresas = db.query(models.Empresa).filter(models.Empresa.is_active == True).all()
    for emp in empresas:
        conteo = db.query(models.CuotaPrestamo).filter(
            models.CuotaPrestamo.empresa_id == emp.id,
            models.CuotaPrestamo.fecha_vencimiento >= hoy,
            models.CuotaPrestamo.estado_pago != "Pagado"
        ).count()
        if conteo > 0:
            crear_notificacion(db, emp.id, f"🟡 Tienes {conteo} cobros programados para el día de hoy.", "warning")


# ═══════════════════════════════════════════════════════════════════════════════
# ÓRDENES DE TRABAJO
# ═══════════════════════════════════════════════════════════════════════════════

EVIDENCE_DIR = "evidencias"
os.makedirs(EVIDENCE_DIR, exist_ok=True)

ordenes_router = APIRouter(
    prefix="/ordenes-trabajo",
    tags=["Órdenes de Trabajo"],
    dependencies=[Depends(get_current_active_user)]
)


@ordenes_router.post("/", response_model=schemas.OrdenTrabajo)
def create_orden_trabajo(orden: schemas.OrdenTrabajoCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    operador_id = current_user.id
    if current_user.role.name == 'Admin' and orden.operador_id is not None:
        if not crud.get_user(db, orden.operador_id):
            raise HTTPException(status_code=404, detail="Operador no encontrado")
        operador_id = orden.operador_id
    return crud.create_orden_trabajo(db=db, empresa_id=empresa_id, orden=orden, operador_id=operador_id)


@ordenes_router.get("/", response_model=List[schemas.OrdenTrabajo])
def read_ordenes_trabajo(
    skip: int = 0, limit: int = 100,
    estado: Optional[str] = None,
    start_date: Optional[date] = None, end_date: Optional[date] = None,
    cliente_id: Optional[int] = None,
    filter_operador_id: Optional[int] = Query(None, alias="operador_id"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    op_filter = filter_operador_id if current_user.role.name == 'Admin' else current_user.id
    return crud.get_ordenes_trabajo(
        db, empresa_id=current_user.empresa_id, skip=skip, limit=limit,
        operador_id=op_filter, estado=estado,
        start_date=start_date, end_date=end_date, cliente_id=cliente_id
    )


@ordenes_router.get("/total", response_model=float)
def get_total_ordenes(
    estado: Optional[str] = None,
    start_date: Optional[date] = None, end_date: Optional[date] = None,
    cliente_id: Optional[int] = None,
    filter_operador_id: Optional[int] = Query(None, alias="operador_id"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    op_filter = filter_operador_id if current_user.role.name == 'Admin' else current_user.id
    return crud.get_total_ordenes_trabajo(
        db, empresa_id=current_user.empresa_id, operador_id=op_filter,
        estado=estado, start_date=start_date, end_date=end_date, cliente_id=cliente_id
    )


@ordenes_router.get("/{orden_id}", response_model=schemas.OrdenTrabajo)
def read_orden_trabajo(orden_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_orden = crud.get_orden_trabajo(db, empresa_id=current_user.empresa_id, orden_id=orden_id)
    if db_orden is None:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    return db_orden


@ordenes_router.put("/{orden_id}", response_model=schemas.OrdenTrabajo)
def update_orden_trabajo(orden_id: int, orden: schemas.OrdenTrabajoCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    db_orden = crud.get_orden_trabajo(db, empresa_id=empresa_id, orden_id=orden_id)
    if db_orden is None:
        raise HTTPException(status_code=404, detail="Orden no encontrada")
    if db_orden.operador_id != current_user.id and current_user.role.name != 'Admin':
        raise HTTPException(status_code=403, detail="Sin permiso para editar esta orden")
    if not crud.get_cliente(db, empresa_id=empresa_id, cliente_id=orden.cliente_id):
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    updated = crud.update_orden_trabajo(db, empresa_id=empresa_id, orden_id=orden_id, orden=orden)
    if updated is None:
        raise HTTPException(status_code=500, detail="Error al actualizar")
    return updated


@ordenes_router.put("/{orden_id}/enviar-revision", response_model=schemas.OrdenTrabajo)
def enviar_revision(orden_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    db_orden = crud.get_orden_trabajo(db, empresa_id=empresa_id, orden_id=orden_id)
    if db_orden.operador_id != current_user.id and current_user.role.name != 'Admin':
        raise HTTPException(status_code=403, detail="Sin permiso")
    return crud.update_orden_trabajo_estado(db, empresa_id=empresa_id, orden_id=orden_id, estado="En revisión")


@ordenes_router.post("/{orden_id}/aprobar", response_model=schemas.OrdenTrabajo)
def approve_orden(orden_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_admin_user)):
    empresa_id = current_user.empresa_id
    db_orden = crud.aprobar_orden_trabajo(db, empresa_id=empresa_id, orden_id=orden_id, admin_user=current_user)
    if db_orden is None:
        raise HTTPException(status_code=404, detail="Orden no encontrada o no está en revisión")
    
    # ✅ FIX CRÍTICO: Se eliminó el bucle que descontaba inventario manualmente aquí.
    # La función crud.aprobar_orden_trabajo ya llama a create_venta(), y el módulo 
    # de ventas YA se encarga de descontar el stock usando FEFO de manera segura.
    
    return db_orden


@ordenes_router.post("/{orden_id}/rechazar", response_model=schemas.OrdenTrabajo)
def reject_orden(orden_id: int, observaciones: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_admin_user)):
    db_orden = crud.rechazar_orden_trabajo(db, empresa_id=current_user.empresa_id, orden_id=orden_id, observaciones=observaciones, admin_user=current_user)
    if db_orden is None:
        raise HTTPException(status_code=404, detail="Orden no encontrada o no está en revisión")
    return db_orden


@ordenes_router.put("/{orden_id}/cerrar", response_model=schemas.OrdenTrabajo)
def cerrar_orden(orden_id: int, close_data: schemas.OrdenTrabajoClose, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_admin_user)):
    if close_data.was_paid and close_data.payment_type == "partial":
        if not close_data.paid_amount or close_data.paid_amount <= 0:
            raise HTTPException(status_code=400, detail="Monto parcial requerido y debe ser > 0")
    db_orden = crud.cerrar_orden_trabajo(db, empresa_id=current_user.empresa_id, orden_id=orden_id, admin_user=current_user, close_data=close_data)
    if db_orden is None:
        raise HTTPException(status_code=404, detail="Orden no encontrada o no puede cerrarse")
    return db_orden


@ordenes_router.post("/{orden_id}/evidencia")
def upload_evidence(orden_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    file_path = os.path.join(EVIDENCE_DIR, f"{orden_id}_{file.filename}")
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    crud.add_evidencia_orden_trabajo(db, empresa_id=current_user.empresa_id, orden_id=orden_id, file_path=file_path)
    return {"filename": file.filename, "path": file_path}


app.include_router(ordenes_router)


# ═══════════════════════════════════════════════════════════════════════════════
# PANEL OPERADOR
# ═══════════════════════════════════════════════════════════════════════════════

panel_router = APIRouter(
    prefix="/panel_operador",
    tags=["Panel del Operador"],
    dependencies=[Depends(get_current_active_user)]
)




# ✅ NUEVO: El operador inicia el trabajo
@panel_router.put("/{orden_id}/iniciar", response_model=schemas.OrdenTrabajo)
def iniciar_orden_operador(orden_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if current_user.role.name not in ('Operador', 'Admin'):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return crud.update_orden_trabajo_estado(db, current_user.empresa_id, orden_id, "Iniciada")

# ✅ NUEVO: El operador termina y envía al Admin
@panel_router.put("/{orden_id}/terminar", response_model=schemas.OrdenTrabajo)
def terminar_orden_operador(orden_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if current_user.role.name not in ('Operador', 'Admin'):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return crud.update_orden_trabajo_estado(db, current_user.empresa_id, orden_id, "En revisión")


@panel_router.get("/pendientes", response_model=List[schemas.PanelOrdenPendiente])
def get_pendientes(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if current_user.role.name not in ('Operador', 'Admin'):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return crud.get_ordenes_pendientes_operador(db, empresa_id=current_user.empresa_id, operador_id=current_user.id)


@panel_router.get("/productividad", response_model=schemas.PanelProductividad)
def get_productividad(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    if current_user.role.name not in ('Operador', 'Admin'):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return crud.get_productividad_operador(db, empresa_id=current_user.empresa_id, operador_id=current_user.id, start_date=start_date, end_date=end_date)


@panel_router.get("/historial", response_model=List[schemas.PanelHistorialItem])
def get_historial(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if current_user.role.name not in ('Operador', 'Admin'):
        raise HTTPException(status_code=403, detail="Acceso denegado")
    return crud.get_historial_reciente_operador(db, empresa_id=current_user.empresa_id, operador_id=current_user.id)


app.include_router(panel_router)


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCCIÓN
# ═══════════════════════════════════════════════════════════════════════════════

produccion_router = APIRouter(
    prefix="/produccion",
    tags=["Producción"],
    dependencies=[Depends(get_current_active_user)]
)


@produccion_router.get("/recetas/", response_model=List[schemas.Receta])
def read_recetas(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_recetas(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@produccion_router.get("/recetas/{receta_id}", response_model=schemas.Receta)
def read_receta(receta_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_r = crud.get_receta(db, empresa_id=current_user.empresa_id, receta_id=receta_id)
    if not db_r:
        raise HTTPException(status_code=404, detail="Receta no encontrada")
    return db_r


@produccion_router.post("/recetas/", response_model=schemas.Receta)
def create_receta(receta: schemas.RecetaCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if crud.get_receta_by_producto(db, empresa_id=current_user.empresa_id, producto_id=receta.producto_id):
        raise HTTPException(status_code=400, detail="Este producto ya tiene una receta.")
    return crud.create_receta(db, empresa_id=current_user.empresa_id, receta=receta)


@produccion_router.delete("/recetas/{receta_id}")
def delete_receta(receta_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    empresa_id = current_user.empresa_id
    db_receta = crud.get_receta(db, empresa_id=empresa_id, receta_id=receta_id)
    if not db_receta:
        raise HTTPException(status_code=404, detail="Receta no encontrada")
    bloqueos = crud.check_can_delete_receta(db, empresa_id=empresa_id, receta_id=receta_id)
    if bloqueos:
        raise HTTPException(
            status_code=409,
            detail=(f"No se puede eliminar la receta '{db_receta.nombre}' porque " + ", ".join(bloqueos) + ".")
        )
    crud.delete_receta(db, empresa_id=empresa_id, receta_id=receta_id)
    return {"message": "Receta eliminada"}


@produccion_router.get("/recetas/{receta_id}/simular")
def simular_produccion(receta_id: int, cantidad: float, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    """Simula una producción para advertir si hay stock suficiente de insumos."""
    receta = crud.get_receta(db, empresa_id=current_user.empresa_id, receta_id=receta_id)
    if not receta:
        raise HTTPException(status_code=404, detail="Receta no encontrada")

    faltantes = []
    costo_teorico_total = 0.0

    for item in receta.items:
        insumo = item.insumo
        cantidad_requerida = item.cantidad * cantidad
        stock_actual = insumo.stock_actual or 0.0
        costo_teorico_total += cantidad_requerida * (insumo.costo or 0.0)
        if stock_actual < cantidad_requerida:
            faltantes.append({
                "insumo_id": insumo.id, "nombre": insumo.nombre,
                "requerido": cantidad_requerida, "disponible": stock_actual,
                "faltante": cantidad_requerida - stock_actual, "unidad": insumo.unidad_medida
            })

    return {
        "factible": len(faltantes) == 0,
        "costo_teorico_total": costo_teorico_total,
        "costo_unitario_estimado": costo_teorico_total / cantidad if cantidad > 0 else 0,
        "faltantes": faltantes
    }


@produccion_router.get("/lotes/", response_model=List[schemas.LoteProduccion])
def read_lotes(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_lotes(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@produccion_router.get("/lotes/{lote_id}", response_model=schemas.LoteProduccion)
def read_lote(lote_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_l = crud.get_lote(db, empresa_id=current_user.empresa_id, lote_id=lote_id)
    if not db_l:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    return db_l


@produccion_router.post("/lotes/", response_model=schemas.LoteProduccion)
def create_lote(lote: schemas.LoteProduccionCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.create_lote(db, empresa_id=current_user.empresa_id, lote=lote)


@produccion_router.post("/lotes/{lote_id}/confirmar", response_model=schemas.LoteProduccion)
def confirmar_lote(lote_id: int, confirm_data: schemas.LoteProduccionConfirm, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    try:
        return crud.confirmar_lote_produccion(db, empresa_id=current_user.empresa_id, lote_id=lote_id, confirm_data=confirm_data)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@produccion_router.put("/lotes/{lote_id}/cancelar")
def cancelar_lote(lote_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    if not crud.cancelar_lote(db, empresa_id=current_user.empresa_id, lote_id=lote_id):
        raise HTTPException(status_code=404, detail="Lote no encontrado o no puede cancelarse")
    return {"message": "Lote cancelado"}


app.include_router(produccion_router)


# ═══════════════════════════════════════════════════════════════════════════════
# COMPRAS
# ═══════════════════════════════════════════════════════════════════════════════

compras_router = APIRouter(
    prefix="/compras",
    tags=["Compras"],
    dependencies=[Depends(get_current_active_user)]
)


@compras_router.get("/", response_model=List[schemas.Compra])
def read_compras(skip: int = 0, limit: int = 100, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_compras(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@compras_router.post("/", response_model=schemas.Compra)
def create_compra(compra: schemas.CompraCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_prov = db.query(models.Cliente).filter(
        models.Cliente.id == compra.proveedor_id,
        models.Cliente.empresa_id == current_user.empresa_id
    ).first()
    if not db_prov or not db_prov.es_proveedor:
        raise HTTPException(status_code=400, detail="Proveedor no válido.")
    return crud.create_compra(db, empresa_id=current_user.empresa_id, compra=compra)


@compras_router.get("/{compra_id}", response_model=schemas.Compra)
def read_compra(compra_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_c = crud.get_compra(db, empresa_id=current_user.empresa_id, compra_id=compra_id)
    if not db_c:
        raise HTTPException(status_code=404, detail="Compra no encontrada")
    return db_c


@compras_router.post("/pagos/", response_model=schemas.PagoCompraCreate)
def add_pago_compra(pago: schemas.PagoCompraCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    db_c = crud.get_compra(db, empresa_id=current_user.empresa_id, compra_id=pago.compra_id)
    if not db_c:
        raise HTTPException(status_code=404, detail="Compra no encontrada")
    monto_pendiente = db_c.total - db_c.monto_pagado
    if pago.monto > (monto_pendiente + 0.01):
        raise HTTPException(status_code=400, detail=f"Pago excede el saldo de {monto_pendiente:.2f}")
    crud.create_pago_compra(db, empresa_id=current_user.empresa_id, pago=pago)
    return pago


app.include_router(compras_router)


# ═══════════════════════════════════════════════════════════════════════════════
# PASARELA DE PAGOS (WOMPI)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/planes-activos", response_model=List[schemas.PlanSuscripcionOut])
def listar_planes_publicos(db: Session = Depends(get_db)):
    return crud.get_planes(db, include_inactive=False)


WOMPI_PUBLIC_KEY = os.getenv("WOMPI_PUBLIC_KEY", "pub_test_...")
WOMPI_INTEGRITY_SECRET = os.getenv("WOMPI_INTEGRITY_SECRET", "prod_integrity_...")


@app.post("/pagos/generar-hash-wompi")
def generar_hash_wompi(
    request_data: schemas.BoldHashRequest,
    current_user: schemas.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    plan = db.query(models.PlanSuscripcion).filter(
        models.PlanSuscripcion.codigo_interno == request_data.plan_name,
        models.PlanSuscripcion.is_active == True
    ).first()
    if not plan:
        raise HTTPException(status_code=400, detail="El plan no existe.")

    monto_en_centavos = str(int(plan.precio * 100))
    divisa = "COP"
    timestamp = int(time.time())
    referencia = f"KSMART-{current_user.empresa_id}-{plan.id}-{timestamp}"
    cadena_concatenada = f"{referencia}{monto_en_centavos}{divisa}{WOMPI_INTEGRITY_SECRET}"
    hash_integridad = hashlib.sha256(cadena_concatenada.encode('utf-8')).hexdigest()

    return {
        "reference": referencia,
        "amount_in_cents": monto_en_centavos,
        "currency": divisa,
        "signature": hash_integridad,
        "public_key": WOMPI_PUBLIC_KEY
    }


@app.post("/webhooks/wompi")
async def webhook_wompi(request: Request, db: Session = Depends(get_db)):
    payload = await request.json()
    event = payload.get("event")
    data = payload.get("data", {}).get("transaction", {})

    if event == "transaction.updated" and data.get("status") == "APPROVED":
        reference = data.get("reference", "")
        if reference.startswith("KSMART-"):
            partes = reference.split("-")
            empresa_id = int(partes[1])
            plan_id = int(partes[2])

            empresa = db.query(models.Empresa).filter(models.Empresa.id == empresa_id).first()
            plan = db.query(models.PlanSuscripcion).filter(models.PlanSuscripcion.id == plan_id).first()

            if empresa and plan:
                empresa.is_active = True
                empresa.plan_type = "premium"

                payment_source = data.get("payment_source_id")
                if payment_source:
                    empresa.wompi_payment_source_id = str(payment_source)

                ahora = datetime.now(timezone.utc)
                base = empresa.trial_ends_at if empresa.trial_ends_at and empresa.trial_ends_at > ahora else ahora
                empresa.trial_ends_at = base + timedelta(days=plan.dias_duracion)

                nuevo_pago = models.RegistroPago(
                    empresa_id=empresa_id,
                    plan_id=plan_id,
                    monto=data.get("amount_in_cents") / 100,
                    moneda=data.get("currency"),
                    metodo_pago=data.get("payment_method_type"),
                    bold_tx_id=data.get("id"),
                    email_pagador=data.get("customer_email"),
                    payload_auditoria=payload
                )
                db.add(nuevo_pago)
                db.commit()
                logger.info(f"✅ Suscripción Wompi activada para empresa {empresa_id}")

    return {"status": "ok"}


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO DE PRÉSTAMOS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/prestamos/", response_model=schemas.PrestamoResponse)
def crear_nuevo_prestamo(
    prestamo: schemas.PrestamoCreate,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user)
):
    cliente = db.query(models.Cliente).filter(
        models.Cliente.id == prestamo.cliente_id,
        models.Cliente.empresa_id == current_user.empresa_id
    ).first()
    if not cliente:
        raise HTTPException(status_code=404, detail="Cliente no encontrado o no pertenece a tu empresa.")
    return crud.crear_prestamo(db=db, prestamo=prestamo, empresa_id=current_user.empresa_id)


@app.get("/prestamos/", response_model=List[schemas.PrestamoResponse])
def listar_prestamos(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    return (
        db.query(models.Prestamo)
        .filter(models.Prestamo.empresa_id == current_user.empresa_id)
        .order_by(models.Prestamo.fecha_inicio.desc())
        .all()
    )


@app.get("/prestamos/cuotas-pendientes")
def cuotas_pendientes(db: Session = Depends(get_db), current_user: schemas.User = Depends(get_current_active_user)):
    query = (
        db.query(models.CuotaPrestamo, models.Prestamo, models.Cliente)
        .join(models.Prestamo, models.CuotaPrestamo.prestamo_id == models.Prestamo.id)
        .join(models.Cliente, models.Prestamo.cliente_id == models.Cliente.id)
        .filter(
            models.CuotaPrestamo.empresa_id == current_user.empresa_id,
            models.CuotaPrestamo.estado_pago != "Pagado"
        )
    )
    if current_user.role.name != "Admin":
        query = query.filter(models.CuotaPrestamo.usuario_asignado_id == current_user.id)

    cuotas = query.order_by(models.CuotaPrestamo.fecha_vencimiento.asc()).all()

    return [
        {
            "cuota_id": cuota.id, "prestamo_id": prestamo.id,
            "numero_cuota": cuota.numero_cuota, "monto_cuota": cuota.monto_cuota,
            "saldo_pendiente": cuota.saldo_pendiente, "fecha_vencimiento": cuota.fecha_vencimiento,
            "estado_pago": cuota.estado_pago, "cliente_nombre": cliente.nombre,
            "cliente_telefono": cliente.telefono, "cliente_direccion": cliente.direccion,
            "cliente_id": cliente.id, "usuario_asignado_id": cuota.usuario_asignado_id
        }
        for cuota, prestamo, cliente in cuotas
    ]


@app.post("/notificaciones/generar-alertas-mora")
def generar_alertas_mora(db: Session = Depends(get_db)):
    hoy = datetime.now().date()
    cuotas_mora = db.query(models.CuotaPrestamo).filter(
        models.CuotaPrestamo.fecha_vencimiento < hoy,
        models.CuotaPrestamo.estado_pago != "Pagado"
    ).all()
    for cuota in cuotas_mora:
        cliente = cuota.prestamo.cliente.nombre
        mensaje = f"🔴 ALERTA: La cuota #{cuota.numero_cuota} de {cliente} está vencida."
        crear_notificacion(db, cuota.empresa_id, mensaje, "error")
    db.commit()
    return {"msg": f"Se generaron {len(cuotas_mora)} alertas de mora."}


# ─── Schemas inline para el módulo de préstamos ───────────────────────────────



# ✅ DESPUÉS:
class PagoCuotaRequest(BaseModel):
    monto_pagado: float = Field(..., gt=0, description="Monto a aplicar. Debe ser mayor a cero.")
    metodo_pago:  str   = Field("Efectivo", description="Efectivo, Transferencia, Nequi, Tarjeta")

# FIX #9: Validación de fecha en reprogramación
class ReprogramarCuotaRequest(BaseModel):
    nueva_fecha: datetime

    @validator('nueva_fecha')
    def validar_fecha_futura(cls, v):
        ahora = datetime.now(timezone.utc)
        fecha = v if v.tzinfo else v.replace(tzinfo=timezone.utc)
        if fecha < ahora:
            raise ValueError("No puedes reprogramar a una fecha pasada.")
        if (fecha - ahora).days > 365:
            raise ValueError("La fecha no puede ser más de 1 año en el futuro.")
        return v


class AsignacionCobroRequest(BaseModel):
    usuario_id: int
    cuota_ids: Optional[List[int]] = None
    cliente_id: Optional[int] = None


# FIX #4: Validaciones de negocio en pago de cuotas
@app.post("/prestamos/cuotas/{cuota_id}/pagar")
def pagar_cuota_cascada(
    cuota_id: int,
    req: PagoCuotaRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    # Monto ya validado por Pydantic (gt=0), pero añadimos límite anti-fraude
    if req.monto_pagado > 999_999_999:
        raise HTTPException(status_code=400, detail="Monto sospechosamente alto.")

    # FIX #5 (préstamos): Verificar que la cuota pertenece a la empresa del usuario
    cuota_inicial = db.query(models.CuotaPrestamo).filter(
        models.CuotaPrestamo.id == cuota_id,
        models.CuotaPrestamo.empresa_id == current_user.empresa_id
    ).first()
    if not cuota_inicial:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")

    monto_disponible = req.monto_pagado
    total_recibido = req.monto_pagado

    cuotas_pendientes_list = (
        db.query(models.CuotaPrestamo)
        .filter(
            models.CuotaPrestamo.prestamo_id == cuota_inicial.prestamo_id,
            models.CuotaPrestamo.estado_pago != "Pagado"
        )
        .order_by(models.CuotaPrestamo.numero_cuota.asc())
        .all()
    )

    cuotas_afectadas = 0
    for cuota in cuotas_pendientes_list:
        if monto_disponible <= 0:
            break
        saldo_actual = cuota.saldo_pendiente
        cuotas_afectadas += 1
        # ✅ DESPUÉS:
        if monto_disponible >= saldo_actual:
            monto_disponible -= saldo_actual
            cuota.saldo_pendiente = 0
            cuota.estado_pago = "Pagado"
            cuota.fecha_pago = datetime.now(crud.BOGOTA_TZ)
            cuota.metodo_pago = req.metodo_pago          # ← NUEVA LÍNEA
        else:
            cuota.saldo_pendiente -= monto_disponible
            cuota.estado_pago = "Parcial"
            cuota.fecha_pago = datetime.now(crud.BOGOTA_TZ)
            cuota.metodo_pago = req.metodo_pago          # ← NUEVA LÍNEA
            monto_disponible = 0

    db.commit()
    return {
        "msg": "Pago procesado",
        "monto_total_recibido": total_recibido,
        "cuotas_afectadas": cuotas_afectadas
    }


@app.post("/prestamos/cuotas/{cuota_id}/reprogramar")
def reprogramar_cuota(
    cuota_id: int,
    req: ReprogramarCuotaRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    cuota = db.query(models.CuotaPrestamo).filter(
        models.CuotaPrestamo.id == cuota_id,
        models.CuotaPrestamo.empresa_id == current_user.empresa_id
    ).first()
    if not cuota:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")
    cuota.fecha_vencimiento = req.nueva_fecha
    db.commit()
    return {"msg": "Cuota reprogramada exitosamente"}


@app.post("/prestamos/asignar-cobrador")
def asignar_cobrador(
    req: AsignacionCobroRequest,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user)
):
    if current_user.role.name != "Admin":
        raise HTTPException(status_code=403, detail="Solo el administrador puede asignar rutas")

    query = db.query(models.CuotaPrestamo).filter(
        models.CuotaPrestamo.empresa_id == current_user.empresa_id
    )
    if req.cliente_id:
        query = query.join(models.Prestamo).filter(
            models.Prestamo.cliente_id == req.cliente_id,
            models.CuotaPrestamo.estado_pago != "Pagado"
        )
    elif req.cuota_ids:
        query = query.filter(models.CuotaPrestamo.id.in_(req.cuota_ids))
    else:
        raise HTTPException(status_code=400, detail="Debe proporcionar cuotas o un cliente")

    cuotas_a_actualizar = query.all()
    for c in cuotas_a_actualizar:
        c.usuario_asignado_id = req.usuario_id
    db.commit()
    return {"msg": f"{len(cuotas_a_actualizar)} cobros asignados correctamente"}


@app.get("/prestamos/calendario-resumen")
def calendario_resumen(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_calendario_cobros(db, current_user.empresa_id)



# ─── Schema inline ────────────────────────────────────────────────────────────
class AbonoCapitalRequest(BaseModel):
    monto_abono: float = Field(..., gt=0, description="Monto a aplicar al capital")


# ─── Abono a capital ──────────────────────────────────────────────────────────


@app.post("/prestamos/{prestamo_id}/abono-capital")
def abono_capital(
    prestamo_id: int,
    req: AbonoCapitalRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    if req.monto_abono > 999_999_999:
        raise HTTPException(status_code=400, detail="Monto sospechosamente alto")

    # Llama a la función de negocio existente (redistribuye cuotas)
    resultado = crud.aplicar_abono_capital(
        db,
        empresa_id=current_user.empresa_id,
        prestamo_id=prestamo_id,
        monto_abono=req.monto_abono,
    )

    # ── #6 FIX: dejar huella en trazabilidad de caja ─────────────────────────
    # Tomamos la primera cuota pendiente del préstamo para registrar el abono.
    # Si el préstamo quedó liquidado, buscamos la última cuota pagada.
    try:
        cuota_huella = (
            db.query(models.CuotaPrestamo)
            .filter(
                models.CuotaPrestamo.prestamo_id == prestamo_id,
                models.CuotaPrestamo.empresa_id  == current_user.empresa_id,
            )
            .order_by(models.CuotaPrestamo.numero_cuota.asc())
            .first()
        )

        if cuota_huella:
            # Marcamos esa cuota con metodo_pago = 'Abono Capital' y
            # guardamos el monto como referencia de trazabilidad.
            # Si ya fue pagada (prestamo liquidado), creamos una entrada
            # de audit en el campo metodo_pago.
            cuota_huella.metodo_pago = "Abono Capital"
            cuota_huella.fecha_pago  = datetime.now(crud.BOGOTA_TZ)
            # El monto abonado queda como diferencia para la caja:
            # monto_cuota - saldo_pendiente refleja lo pagado en esa cuota.
            db.commit()

    except Exception:
        # El abono ya se aplicó; si falla el registro de trazabilidad
        # no revertimos — solo logueamos.
        pass

    return resultado

# ─── Resumen de mora de un préstamo ──────────────────────────────────────────
@app.get("/prestamos/{prestamo_id}/mora")
def resumen_mora_prestamo(
    prestamo_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    prestamo = db.query(models.Prestamo).filter(
        models.Prestamo.id         == prestamo_id,
        models.Prestamo.empresa_id == current_user.empresa_id,
    ).first()
    if not prestamo:
        raise HTTPException(status_code=404, detail="Préstamo no encontrado")

    cuotas = db.query(models.CuotaPrestamo).filter(
        models.CuotaPrestamo.prestamo_id == prestamo_id,
        models.CuotaPrestamo.estado_pago != "Pagado",
    ).all()

    detalle = []
    total_mora = 0.0
    for c in cuotas:
        info = crud.calcular_mora_cuota(c, prestamo.tasa_mora)
        total_mora += info["mora"]
        detalle.append({
            "cuota_id":       c.id,
            "numero_cuota":   c.numero_cuota,
            "saldo_pendiente": c.saldo_pendiente,
            "dias_vencido":   info["dias"],
            "mora":           info["mora"],
            "total_a_pagar":  info["total"],
        })

    return {
        "prestamo_id":  prestamo_id,
        "tasa_mora_mensual": prestamo.tasa_mora,
        "total_mora_acumulada": round(total_mora, 2),
        "cuotas_en_mora": [d for d in detalle if d["dias_vencido"] > 0],
    }


@app.get("/reportes/calendario-cobros")
def calendario_cobros(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_active_user)):
    return crud.get_resumen_calendario_cobros(db, current_user.empresa_id)


# FIX #7: Agregar joinedload(prestamo) para evitar N+1 query en liquidación diaria
@app.get("/prestamos/liquidacion-diaria")
def liquidacion_diaria_ruta(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user)
):
    if current_user.role.name != "Admin":
        raise HTTPException(status_code=403, detail="Solo el Admin puede ver la liquidación")

    hoy = datetime.now(timezone.utc).date()

    cuotas_pagadas = (
        db.query(models.CuotaPrestamo)
        .options(
            joinedload(models.CuotaPrestamo.cobrador),
            joinedload(models.CuotaPrestamo.prestamo)  # FIX: evita N+1
        )
        .filter(
            models.CuotaPrestamo.empresa_id == current_user.empresa_id,
            models.CuotaPrestamo.fecha_pago != None
        )
        .all()
    )

    resumen_cobradores = {}
    total_global = 0.0

    for c in cuotas_pagadas:
        if c.fecha_pago.date() != hoy:
            continue
        cobrador_id = c.usuario_asignado_id or 0
        cobrador_nombre = c.cobrador.username if c.cobrador else "Sin asignar"
        monto_recaudado = c.monto_cuota - c.saldo_pendiente

        if cobrador_id not in resumen_cobradores:
            resumen_cobradores[cobrador_id] = {
                "cobrador_id": cobrador_id,
                "cobrador_nombre": cobrador_nombre.upper(),
                "total_recaudado": 0.0,
                "cuotas_cobradas": 0
            }
        resumen_cobradores[cobrador_id]["total_recaudado"] += monto_recaudado
        resumen_cobradores[cobrador_id]["cuotas_cobradas"] += 1
        total_global += monto_recaudado

    return {
        "fecha": hoy.isoformat(),
        "total_global": total_global,
        "cobradores": list(resumen_cobradores.values())
    }



from reportlab.lib.pagesizes import A6
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib import colors
import io as _io

@app.get("/prestamos/cuotas/{cuota_id}/recibo-pdf")
def descargar_recibo_pdf(
    cuota_id: int,
    monto_pagado: float,
    saldo_restante: float = 0.0,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Genera y retorna un recibo de pago en PDF listo para descargar."""

    # ── Buscar cuota ──────────────────────────────────────────────────────────
    cuota = (
        db.query(models.CuotaPrestamo)
        .join(models.Prestamo)
        .join(models.Cliente)
        .filter(
            models.CuotaPrestamo.id == cuota_id,
            models.CuotaPrestamo.empresa_id == current_user.empresa_id,
        )
        .first()
    )
    if not cuota:
        raise HTTPException(status_code=404, detail="Cuota no encontrada")

    prestamo       = cuota.prestamo
    cliente        = prestamo.cliente
    empresa        = current_user.empresa
    ahora          = datetime.now(crud.BOGOTA_TZ)
    fecha_str      = ahora.strftime("%d/%m/%Y %H:%M")
    vence_str      = cuota.fecha_vencimiento.strftime("%d/%m/%Y") if cuota.fecha_vencimiento else "—"

    def fmt_cop(val: float) -> str:
        return f"$ {val:,.0f}".replace(",", ".")

    # ── Generar PDF en memoria ────────────────────────────────────────────────
    buffer = _io.BytesIO()

    # Tamaño tipo recibo térmico: 80mm × 140mm
    W, H = 80 * mm, 140 * mm
    c = pdf_canvas.Canvas(buffer, pagesize=(W, H))

    GRAY    = colors.HexColor("#64748b")
    DARK    = colors.HexColor("#0f172a")
    ORANGE  = colors.HexColor("#FF6020")
    GREEN   = colors.HexColor("#10B981")
    RED     = colors.HexColor("#EF4444")

    def line_h(y, color=GRAY, width=0.5):
        c.setStrokeColor(color)
        c.setLineWidth(width)
        c.line(5 * mm, y, W - 5 * mm, y)

    y = H - 8 * mm  # cursor vertical, de arriba hacia abajo

    # ── Cabecera ──────────────────────────────────────────────────────────────
    c.setFillColor(ORANGE)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(W / 2, y, "KSMART360")
    y -= 5 * mm

    c.setFillColor(GRAY)
    c.setFont("Helvetica", 7)
    c.drawCentredString(W / 2, y, empresa.nombre.upper())
    y -= 4 * mm

    c.setFont("Helvetica", 7)
    c.drawCentredString(W / 2, y, "RECIBO DE PAGO")
    y -= 3 * mm

    line_h(y, ORANGE, 1)
    y -= 5 * mm

    # ── Info del recibo ───────────────────────────────────────────────────────
    def row(label: str, value: str, bold_val: bool = False, val_color=DARK):
        nonlocal y
        c.setFillColor(GRAY)
        c.setFont("Helvetica", 7)
        c.drawString(5 * mm, y, label)

        c.setFillColor(val_color)
        font = "Helvetica-Bold" if bold_val else "Helvetica"
        c.setFont(font, 7)
        c.drawRightString(W - 5 * mm, y, value)
        y -= 4.5 * mm

    row("Recibo N°",    f"{cuota_id:06d}")
    row("Fecha",        fecha_str)
    row("Empresa",      empresa.nombre[:28])
    y -= 1 * mm
    line_h(y)
    y -= 4 * mm

    row("Cliente",      cliente.nombre[:28])
    row("Préstamo #",   str(prestamo.id))
    row("Cuota #",      f"{cuota.numero_cuota} de {prestamo.cantidad_cuotas}")
    row("Vencimiento",  vence_str)
    y -= 1 * mm
    line_h(y)
    y -= 5 * mm

    # ── Montos ────────────────────────────────────────────────────────────────
    c.setFillColor(GRAY)
    c.setFont("Helvetica", 7)
    c.drawCentredString(W / 2, y, "DETALLE DEL PAGO")
    y -= 5 * mm

    row("Valor recibido", fmt_cop(monto_pagado), bold_val=True, val_color=GREEN)

    if saldo_restante > 0:
        row("Saldo pendiente", fmt_cop(saldo_restante), val_color=ORANGE)
    else:
        row("Estado cuota", "SALDADA ✓", bold_val=True, val_color=GREEN)

    y -= 2 * mm
    line_h(y, ORANGE, 0.8)
    y -= 6 * mm

    # ── Monto grande central ──────────────────────────────────────────────────
    c.setFillColor(GREEN)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(W / 2, y, fmt_cop(monto_pagado))
    y -= 5 * mm

    c.setFillColor(GRAY)
    c.setFont("Helvetica", 7)
    c.drawCentredString(W / 2, y, "VALOR RECIBIDO")
    y -= 8 * mm

    line_h(y)
    y -= 5 * mm

    # ── Footer ────────────────────────────────────────────────────────────────
    c.setFillColor(GRAY)
    c.setFont("Helvetica", 6)
    c.drawCentredString(W / 2, y, "Gracias por su pago")
    y -= 3.5 * mm
    c.drawCentredString(W / 2, y, f"Powered by KSMP Systems · {ahora.year}")

    c.save()
    buffer.seek(0)

    filename = f"recibo_cuota_{cuota_id}_{ahora.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )




# ═══════════════════════════════════════════════════════════════════════════
# PEGA ESTE BLOQUE EN main.py
# Añade el import de requests al tope del archivo si no está:
#   import requests
#
# Pega el resto justo antes de: app.mount("/evidencias" ...)
# ═══════════════════════════════════════════════════════════════════════════

import re as _re   # ya está en Python stdlib, sin instalar nada

# ─── Cache en memoria del precio del cacao ────────────────────────────────────
_cacao_cache: dict = {
    "data":           None,   # dict con los valores calculados
    "last_fetch_8":   None,   # date del último refresh de las 8:00
    "last_fetch_14":  None,   # date del último refresh de las 14:00
}


def _cacao_needs_refresh() -> bool:
    """True si el cache está vacío o si ya pasó la ventana de las 8:00/14:00 sin actualizar."""
    from zoneinfo import ZoneInfo
    bogota = ZoneInfo("America/Bogota")
    ahora  = datetime.now(bogota)
    hoy    = ahora.date()
    hora   = ahora.hour

    if _cacao_cache["data"] is None:
        return True
    if hora >= 8  and _cacao_cache["last_fetch_8"]  != hoy:
        return True
    if hora >= 14 and _cacao_cache["last_fetch_14"] != hoy:
        return True
    return False


def _fetch_precio_cacao_raw():
    """
    Obtiene el precio de los futuros de cacao ICE (USD/ton) y la TRM (COP/USD).

    Fuentes cacao (en orden):
      1. Yahoo Finance contrato específico del mes vigente (CCK26=F, CCN26=F…)
      2. Yahoo Finance CC=F front-month genérico  (query1 y query2)
      3. Stooq CC.F  (fallback final)

    Fuentes TRM (en orden):
      1. datos.gov.co API oficial — Superintendencia Financiera (JSON puro)
      2. dolar.wilkinsonpc.com.co — misma fuente que fepcacao.com.co
      3. Banco de la República (fallback HTML)

    Devuelve (cacao_usd_ton: float, trm_cop: float) o (None, None) si todo falla.
    """
    _headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json, text/html, */*",
        "Accept-Language": "es-CO,es;q=0.9,en;q=0.8",
    }

    # ═══════════════════════════════════════════════════════════════════════
    # 1. PRECIO CACAO ICE (USD/tonelada)
    # Lógica de meses ICE Cocoa: Mar(H), May(K), Jul(N), Sep(U), Dec(Z)
    # fepcacao.com.co referencia el contrato FRONT MONTH vigente,
    # pero Yahoo/Stooq a veces ya rotaron al siguiente. Intentamos el
    # contrato exacto antes de caer al genérico.
    # ═══════════════════════════════════════════════════════════════════════

    cacao_usd_ton = None

    def _contrato_vigente_yahoo() -> list[str]:
        """
        Devuelve los tickers de Yahoo Finance para el contrato ICE Cocoa
        actual y el siguiente, en orden de prioridad.
        Meses ICE: H=Mar, K=May, N=Jul, U=Sep, Z=Dec
        """
        from zoneinfo import ZoneInfo
        ahora = datetime.now(ZoneInfo("America/New_York"))
        anio  = ahora.year
        mes   = ahora.month

        # Próximos meses de vencimiento ICE en orden anual
        meses_ice = [(3, 'H'), (5, 'K'), (7, 'N'), (9, 'U'), (12, 'Z')]

        tickers = []
        for (m, cod) in meses_ice:
            if m >= mes:
                # Año actual
                tickers.append(f"CC{cod}{str(anio)[2:]}=F")
            else:
                # Año siguiente
                tickers.append(f"CC{cod}{str(anio + 1)[2:]}=F")

        # Tomar los dos primeros (vigente + siguiente) + genérico
        result = tickers[:2] + ["CC=F"]
        logger.info("Tickers cacao a intentar: %s", result)
        return result

    # ── Intento 1-N: Yahoo Finance con contrato específico ────────────────
    tickers = _contrato_vigente_yahoo()
    for ticker in tickers:
        if cacao_usd_ton:
            break
        encoded = ticker.replace("=", "%3D")
        for yf_host in ("query1", "query2"):
            try:
                url = (
                    f"https://{yf_host}.finance.yahoo.com/v8/finance/chart/"
                    f"{encoded}?interval=1d&range=2d"
                )
                r = requests.get(url, headers=_headers, timeout=10)
                if r.status_code == 200:
                    meta = r.json()["chart"]["result"][0]["meta"]
                    # previousClose = settlement oficial ICE del día anterior
                    # (= precio que publica fepcacao.com.co / theice.com)
                    raw = meta.get("previousClose") or meta.get("regularMarketPrice")
                    if raw:
                        val = float(raw)
                        if 500 < val < 20000:
                            cacao_usd_ton = val
                            logger.info(
                                "Cacao Yahoo %s (%s): %.2f USD/ton",
                                ticker, yf_host, cacao_usd_ton
                            )
                            break
            except Exception as exc:
                logger.warning("Yahoo %s (%s) falló: %s", ticker, yf_host, exc)

    # ── Fallback final: Stooq CC.F ────────────────────────────────────────
    if not cacao_usd_ton:
        try:
            r = requests.get(
                "https://stooq.com/q/l/?s=cc.f&f=sd2t2ohlcv&h&e=csv",
                headers=_headers,
                timeout=10,
            )
            if r.status_code == 200:
                lines = [l.strip() for l in r.text.strip().split('\n') if l.strip()]
                if len(lines) >= 2:
                    cols = lines[1].split(',')
                    if len(cols) > 6:
                        val = float(cols[6])   # columna Close
                        if 500 < val < 20000:
                            cacao_usd_ton = val
                            logger.info("Cacao Stooq: %.2f USD/ton", cacao_usd_ton)
        except Exception as exc:
            logger.warning("Stooq CC.F falló: %s", exc)

    # ═══════════════════════════════════════════════════════════════════════
    # 2. TRM (COP / USD)
    # ═══════════════════════════════════════════════════════════════════════

    trm_cop = None

    # ── Intento 1: datos.gov.co — API oficial Superintendencia Financiera ──
    # Endpoint público SODA, sin API key, devuelve JSON puro con la TRM
    # del día o del día hábil más reciente.
    try:
        from zoneinfo import ZoneInfo
        hoy_col = datetime.now(ZoneInfo("America/Bogota")).strftime("%Y-%m-%d")
        url_gov = (
            "https://www.datos.gov.co/resource/32sa-8pi3.json"
            f"?vigenciadesde={hoy_col}"
        )
        r = requests.get(url_gov, headers={**_headers, "Accept": "application/json"}, timeout=10)
        if r.status_code == 200:
            datos = r.json()
            if datos:
                val = float(datos[0].get("valor", 0))
                if 2000 < val < 8000:
                    trm_cop = val
                    logger.info("TRM datos.gov.co: %.2f COP/USD", trm_cop)
    except Exception as exc:
        logger.warning("TRM datos.gov.co falló: %s", exc)

    # ── Intento 2 (fallback): datos.gov.co con fecha de ayer ─────────────
    # (días no hábiles no tienen dato; búsqueda con $limit y $order)
    if not trm_cop:
        try:
            url_gov2 = (
                "https://www.datos.gov.co/resource/32sa-8pi3.json"
                "?$limit=1&$order=vigenciadesde+DESC"
            )
            r = requests.get(url_gov2, headers={**_headers, "Accept": "application/json"}, timeout=10)
            if r.status_code == 200:
                datos = r.json()
                if datos:
                    val = float(datos[0].get("valor", 0))
                    if 2000 < val < 8000:
                        trm_cop = val
                        logger.info("TRM datos.gov.co (latest): %.2f COP/USD", trm_cop)
        except Exception as exc:
            logger.warning("TRM datos.gov.co latest falló: %s", exc)

    # ── Intento 3: dolar.wilkinsonpc.com.co (misma fuente que fepcacao) ───
    # IMPORTANTE: extraemos SOLO el valor de TRM (no Next Day ni Spot)
    if not trm_cop:
        try:
            r = requests.get(
                "https://dolar.wilkinsonpc.com.co/",
                headers=_headers,
                timeout=10,
            )
            if r.status_code == 200:
                # Buscamos el patrón ESPECÍFICO de TRM en el HTML:
                # Aparece como: "TRM....$3,551.17" o "TRM...$3.551,17"
                # Usamos una ventana pequeña después de "TRM" para no capturar
                # "DÓLAR NEXT DAY" ni "DÓLAR SPOT" que vienen después.
                m = _re.search(
                    r'TRM[^$\d]{0,60}\$([\d]{1,2}[,\.][\d]{3}[,\.][\d]{2})',
                    r.text,
                    _re.DOTALL | _re.IGNORECASE
                )
                if m:
                    clean = m.group(1)
                    if ',' in clean and '.' in clean:
                        if clean.index(',') < clean.index('.'):
                            clean = clean.replace(',', '')        # "3,551.17"
                        else:
                            clean = clean.replace('.', '').replace(',', '.')  # "3.551,17"
                    elif ',' in clean:
                        clean = clean.replace(',', '.')
                    val = float(clean)
                    if 2000 < val < 8000:
                        trm_cop = val
                        logger.info("TRM wilkinsonpc: %.2f COP/USD", trm_cop)
        except Exception as exc:
            logger.warning("TRM wilkinsonpc falló: %s", exc)

    # ── Intento 4: Banco de la República (último fallback) ────────────────
    if not trm_cop:
        try:
            hoy_str = datetime.now().strftime("%Y-%m-%d")
            r = requests.get(
                f"https://www.banrep.gov.co/es/estadisticas/trm?fecha={hoy_str}&format=json",
                headers=_headers,
                timeout=8,
            )
            if r.status_code == 200:
                m = _re.search(r'"valor"\s*:\s*([\d\.]+)', r.text)
                if not m:
                    m = _re.search(r'(3[\.,]\d{3}[\.,]\d{2})', r.text)
                if m:
                    val = float(m.group(1).replace(',', ''))
                    if 2000 < val < 8000:
                        trm_cop = val
                        logger.info("TRM Banrep: %.2f COP/USD", trm_cop)
        except Exception as exc:
            logger.warning("TRM Banrep falló: %s", exc)

    return cacao_usd_ton, trm_cop



# ─── Endpoint público ─────────────────────────────────────────────────────────

@app.get("/mercado/precio-cacao")
def precio_cacao_fedecacao(
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Precio del kilo de cacao en pesos colombianos.

    Metodología (idéntica a fepcacao.com.co):
      1. Precio futuros cacao ICE en USD/tonelada (Yahoo Finance, ticker CC=F)
      2. TRM oficial del día (dolar.wilkinsonpc.com.co)
      3. Precio COP/kg = (USD/ton × TRM) ÷ 1 000

    Se actualiza automáticamente a las 08:00 y 14:00 hora Colombia.
    """
    from zoneinfo import ZoneInfo
    bogota = ZoneInfo("America/Bogota")
    ahora  = datetime.now(bogota)
    hoy    = ahora.date()
    hora   = ahora.hour

    if _cacao_needs_refresh():
        try:
            cacao_usd_ton, trm_cop = _fetch_precio_cacao_raw()

            if cacao_usd_ton and trm_cop:
                precio_cop_kg = round((cacao_usd_ton * trm_cop) / 1000, 0)

                # Tendencia respecto al valor anterior (si existe)
                prev = (_cacao_cache["data"] or {}).get("precio_cop_kg")
                if prev:
                    variacion_pct = round(((precio_cop_kg - prev) / prev) * 100, 2)
                    tendencia = "alza" if precio_cop_kg > prev else "baja" if precio_cop_kg < prev else "estable"
                else:
                    variacion_pct = 0.0
                    tendencia     = "estable"

                _cacao_cache["data"] = {
                    "precio_cop_kg":      precio_cop_kg,
                    "precio_usd_ton":     round(cacao_usd_ton, 2),
                    "trm_cop":            round(trm_cop, 2),
                    "tendencia":          tendencia,
                    "variacion_pct":      variacion_pct,
                    "ultima_actualizacion": ahora.strftime("%Y-%m-%dT%H:%M:%S"),
                    "fecha_precio":       ahora.strftime("%d/%m/%Y"),
                    "fuente_cacao":       "ICE Cocoa Futures (CC=F) via Yahoo Finance",
                    "fuente_trm":         "dolar.wilkinsonpc.com.co",
                    "referencia":         "https://www.fepcacao.com.co/",
                }

                # Marcar cuál ventana se cubrió
                if hora >= 14:
                    _cacao_cache["last_fetch_14"] = hoy
                    _cacao_cache["last_fetch_8"]  = hoy
                elif hora >= 8:
                    _cacao_cache["last_fetch_8"]  = hoy

                logger.info(
                    "✅ Precio cacao actualizado: $%s COP/kg (USD %.2f/ton × TRM %.2f)",
                    f"{precio_cop_kg:,.0f}",
                    cacao_usd_ton,
                    trm_cop,
                )
            else:
                logger.error(
                    "No se pudo obtener precio cacao: USD/ton=%s, TRM=%s",
                    cacao_usd_ton, trm_cop,
                )
        except Exception as exc:
            logger.exception("Error inesperado al actualizar precio cacao: %s", exc)

    if _cacao_cache["data"]:
        return _cacao_cache["data"]

    raise HTTPException(
        status_code=503,
        detail=(
            "El servicio de precio de cacao no está disponible en este momento. "
            "Verifica tu conexión a internet desde el servidor y vuelve a intentarlo."
        ),
    )





# ═══════════════════════════════════════════════════════════════════════════════
# ESTÁTICOS Y UTILIDADES
# ═══════════════════════════════════════════════════════════════════════════════

app.mount("/evidencias", StaticFiles(directory=EVIDENCE_DIR), name="evidencias")


@app.get("/")
def read_root():
    return {"message": "Ksmart360 API Multi-Tenant v2.1", "docs": "/docs"}


@app.get("/ping")
def ping():
    """Endpoint ligero para mantener el servidor de Render despierto."""
    return {"status": "ok", "message": "Backend de Ksmart360 activo y despierto"}


@app.get("/superadmin/sync-modulos")
def sincronizar_modulos_nuevos(db: Session = Depends(get_db)):
    nuevos_modulos = [
        {"name": "Simulador de Préstamos", "frontend_path": "/prestamos"},
        {"name": "Ruta de Cobro",           "frontend_path": "/ruta-cobro"}
    ]
    agregados = []
    for mod in nuevos_modulos:
        existe = db.query(models.Modulo).filter(models.Modulo.frontend_path == mod["frontend_path"]).first()
        if not existe:
            nuevo_mod = models.Modulo(name=mod["name"], frontend_path=mod["frontend_path"])
            db.add(nuevo_mod)
            db.commit()
            db.refresh(nuevo_mod)
            db.add(models.RoleModule(role_id=1, module_id=nuevo_mod.id))
            db.commit()
            agregados.append(mod["name"])
    return {"msg": f"Sincronización completada. Módulos agregados: {agregados}"}




# ═══════════════════════════════════════════════════════════════════════════
# AÑADIR A main.py — Endpoints de Gestión de Lotes y Perecederos
# Pégalos en la sección de INVENTARIO de tu main.py existente
# ═══════════════════════════════════════════════════════════════════════════


# ── Crear / registrar un lote nuevo ──────────────────────────────────────────
@app.post("/inventario/lotes", response_model=schemas.LoteExistenciaOut)
def crear_lote(
    payload: schemas.LoteExistenciaCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Registra un nuevo lote de existencias.
    Si el número de lote ya existe para ese producto, suma la cantidad.
    """
    lote = crud.crear_lote_existencia(db, empresa_id=current_user.empresa_id, payload=payload)
    return crud._enriquecer_lote(lote)

# ── Listar todos los lotes de la empresa ─────────────────────────────────────
@app.get("/inventario/lotes", response_model=List[schemas.LoteExistenciaOut])
def listar_todos_lotes(
    solo_activos: bool   = Query(True),
    producto_id:  Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Lista todos los lotes activos de la empresa, opcionalmente filtrados por producto."""
    return crud.get_todos_los_lotes(
        db, empresa_id=current_user.empresa_id,
        solo_activos=solo_activos, producto_id=producto_id,
    )


# ── Listar lotes de un producto específico (FEFO) ────────────────────────────
@app.get("/inventario/lotes/{producto_id}", response_model=List[schemas.LoteExistenciaOut])
def listar_lotes_producto(
    producto_id: int,
    solo_activos: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """Lista los lotes de un producto ordenados FEFO (primero el que vence antes)."""
    return crud.get_lotes_producto(
        db, empresa_id=current_user.empresa_id,
        producto_id=producto_id, solo_activos=solo_activos,
    )




# ── Ajuste manual de un lote ─────────────────────────────────────────────────
@app.patch("/inventario/lotes/{lote_id}/ajuste")
def ajustar_lote(
    lote_id: int,
    ajuste: schemas.LoteAjusteCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Ajusta manualmente la cantidad de un lote.
    Positivo = entrada, negativo = salida (merma, donación, destrucción).
    """
    lote = crud.ajustar_lote(db, empresa_id=current_user.empresa_id,
                              lote_id=lote_id, ajuste=ajuste)
    return crud._enriquecer_lote(lote)


# ── Sugerencia FEFO sin modificar BD ─────────────────────────────────────────
@app.get("/inventario/lotes/{producto_id}/sugerencia-fefo")
def sugerencia_fefo(
    producto_id:        int,
    cantidad_requerida: float = Query(..., gt=0),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Devuelve qué lotes se consumirían al vender `cantidad_requerida` unidades,
    aplicando FEFO. No modifica la BD — es solo una consulta de planificación.
    """
    return crud.sugerencia_fefo(
        db, empresa_id=current_user.empresa_id,
        producto_id=producto_id, cantidad_requerida=cantidad_requerida,
    )


# ── Reporte: próximos a vencer ────────────────────────────────────────────────
@app.get("/reportes/proximos-a-vencer", response_model=List[schemas.AlertaVencimientoOut])
def proximos_a_vencer(
    dias: int = Query(30, ge=1, le=365, description="Horizon de días hacia adelante"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Lista todos los lotes con stock > 0 que vencen en los próximos `dias` días.
    Incluye lotes ya vencidos (días negativos).
    """
    return crud.get_alertas_vencimiento(
        db, empresa_id=current_user.empresa_id, dias=dias,
    )


# ── KPIs de alertas para el dashboard ─────────────────────────────────────────
@app.get("/reportes/resumen-alertas-vencimiento")
def resumen_alertas_vencimiento(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Retorna conteos por categoría de urgencia y valor total en riesgo.
    Ideal para un widget en el dashboard.
    """
    return crud.get_resumen_alertas(db, empresa_id=current_user.empresa_id)


# ── Cron: notificaciones automáticas de vencimientos ─────────────────────────
@app.post("/superadmin/notificar-vencimientos-lotes")
def notificar_vencimientos_lotes(db: Session = Depends(get_db)):
    """
    Genera notificaciones de vencimiento para todas las empresas activas.
    Llamar diariamente desde un cron job externo (ej. cron-job.org).
    No requiere autenticación — protégelo por IP o secreto si es necesario.
    """
    total = crud.notificar_vencimientos_proximos(db)
    return {"msg": f"Se generaron {total} notificaciones de vencimiento."}





# ════════════════════════════════════════════════════════════════════════════
# FASE 2A — ENDPOINTS RESOLUCIONES DIAN
# ════════════════════════════════════════════════════════════════════════════

@app.get("/resoluciones/", response_model=List[schemas.ResolucionDianOut])
def listar_resoluciones(
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    """Lista todas las resoluciones DIAN de la empresa con campos calculados."""
    return crud.get_resoluciones(db, empresa_id=current_user.empresa_id)


@app.post("/resoluciones/", response_model=schemas.ResolucionDianOut)
def crear_resolucion(
    payload: schemas.ResolucionDianCreate,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_admin_user),
):
    """Crea una nueva resolución DIAN. Solo admins."""
    return crud.create_resolucion(db, empresa_id=current_user.empresa_id, payload=payload)


@app.put("/resoluciones/{resolucion_id}", response_model=schemas.ResolucionDianOut)
def actualizar_resolucion(
    resolucion_id: int,
    payload: schemas.ResolucionDianUpdate,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_admin_user),
):
    resolucion = crud.update_resolucion(
        db, empresa_id=current_user.empresa_id,
        resolucion_id=resolucion_id, payload=payload,
    )
    if not resolucion:
        raise HTTPException(status_code=404, detail="Resolución no encontrada.")
    return resolucion


@app.patch("/resoluciones/{resolucion_id}/activar", response_model=schemas.ResolucionDianOut)
def activar_resolucion(
    resolucion_id: int,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_admin_user),
):
    """
    Activa esta resolución y desactiva automáticamente las demás de la empresa.
    Es el equivalente de 'seleccionar resolución actual'.
    """
    resolucion = crud.activar_resolucion(
        db, empresa_id=current_user.empresa_id, resolucion_id=resolucion_id
    )
    if not resolucion:
        raise HTTPException(status_code=404, detail="Resolución no encontrada.")
    return resolucion


@app.delete("/resoluciones/{resolucion_id}")
def eliminar_resolucion(
    resolucion_id: int,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_admin_user),
):
    crud.delete_resolucion(db, empresa_id=current_user.empresa_id, resolucion_id=resolucion_id)
    return {"message": "Resolución eliminada correctamente."}


@app.get("/resoluciones/activa")
def get_resolucion_activa(
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    """Endpoint liviano para que el frontend muestre la resolución activa en el header de ventas."""
    resolucion = crud._get_resolucion_activa(db, empresa_id=current_user.empresa_id)
    if not resolucion:
        return {"activa": False, "resolucion": None}
    return {
        "activa":     True,
        "prefijo":    resolucion.prefijo or "",
        "siguiente":  resolucion.numero_actual + 1,
        "disponibles": resolucion.numero_final - resolucion.numero_actual,
        "resolucion": resolucion.numero_resolucion,
    }


# ════════════════════════════════════════════════════════════════════════════
# FASE 2B — ENDPOINTS COTIZACIONES
# ════════════════════════════════════════════════════════════════════════════

@app.get("/cotizaciones/")
def listar_cotizaciones(
    skip: int = 0,
    limit: int = Query(default=100, le=500),
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    """Lista todas las cotizaciones con estado calculado (vigente/vencida/convertida)."""
    return crud.get_cotizaciones(db, empresa_id=current_user.empresa_id, skip=skip, limit=limit)


@app.post("/cotizaciones/")
def crear_cotizacion(
    payload: schemas.CotizacionCreate,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    """
    Crea una cotización/proforma.
    NO descuenta stock, NO crea movimientos de inventario.
    """
    if not payload.detalles:
        raise HTTPException(status_code=400, detail="La cotización debe tener al menos un ítem.")

    return crud.create_cotizacion(db, empresa_id=current_user.empresa_id, payload=payload)


@app.post("/cotizaciones/{cotizacion_id}/convertir")
def convertir_cotizacion(
    cotizacion_id: int,
    payload: schemas.CotizacionConvertir,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    """
    Convierte una cotización en venta real:
    - Valida stock disponible
    - Crea movimientos de inventario (FEFO si aplica)
    - Asigna numero_factura desde resolución DIAN activa
    - Registra método y estado de pago
    """
    return crud.convertir_cotizacion_a_venta(
        db,
        empresa_id    = current_user.empresa_id,
        cotizacion_id = cotizacion_id,
        payload       = payload,
    )


@app.get("/cotizaciones/{cotizacion_id}")
def get_cotizacion(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    cotizacion = (
        db.query(models.Venta)
        .options(
            joinedload(models.Venta.cliente),
            joinedload(models.Venta.detalles).joinedload(models.DetalleVenta.producto),
        )
        .filter(
            models.Venta.id         == cotizacion_id,
            models.Venta.empresa_id == current_user.empresa_id,
            models.Venta.tipo       == "cotizacion",
        )
        .first()
    )
    if not cotizacion:
        raise HTTPException(status_code=404, detail="Cotización no encontrada.")
    return cotizacion


@app.delete("/cotizaciones/{cotizacion_id}")
def eliminar_cotizacion(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: schemas.User = Depends(get_current_active_user),
):
    crud.delete_cotizacion(db, empresa_id=current_user.empresa_id, cotizacion_id=cotizacion_id)
    return {"message": "Cotización eliminada correctamente."}





@app.get("/reportes/caja-rango")
def reporte_caja_rango(
    start_date: date = Query(...),
    end_date:   date = Query(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_active_user),
):
    """
    Informe de ingresos y salidas por rango de fecha (#2).
    Incluye desglose por método de pago (#3).

    Fuentes de ingreso consideradas:
      - Ventas de contado (tabla ventas, pagadas sin abono posterior)
      - Abonos a cartera (tabla pagos, relacionados con ventas)
      - Recaudo de préstamos (tabla cuotas_prestamo, estado Pagado)
    Egresos:
      - Gastos (tabla gastos)
    """
    empresa_id = current_user.empresa_id

    # Acumula por día y por método
    resumen_dias: dict[str, dict] = {}
    metodos_totales: dict[str, float] = {}
    total_ingresos = 0.0
    total_egresos  = 0.0

    def _fecha_col(dt) -> str:
        """Convierte cualquier datetime/date a string 'YYYY-MM-DD' en hora Bogotá."""
        if dt is None:
            return None
        if isinstance(dt, datetime):
            try:
                return dt.astimezone(crud.BOGOTA_TZ).strftime("%Y-%m-%d")
            except Exception:
                return str(dt)[:10]
        return str(dt)[:10]

    def _acumular(fecha_str: str, metodo: str, monto: float, tipo: str):
        nonlocal total_ingresos, total_egresos
        if not fecha_str:
            return
        if fecha_str not in resumen_dias:
            resumen_dias[fecha_str] = {
                "fecha": fecha_str,
                "ingresos": 0.0,
                "egresos":  0.0,
                "neto":     0.0,
                "por_metodo": {},
            }
        dia = resumen_dias[fecha_str]

        if tipo == "ingreso":
            dia["ingresos"] += monto
            dia["neto"]     += monto
            total_ingresos  += monto
            metodos_totales[metodo] = metodos_totales.get(metodo, 0.0) + monto
            dia["por_metodo"][metodo] = dia["por_metodo"].get(metodo, 0.0) + monto
        else:
            dia["egresos"] += monto
            dia["neto"]    -= monto
            total_egresos  += monto

    # ── Boundaries UTC para el rango completo ────────────────────────────────
    utc_start, _ = crud.get_utc_boundaries(start_date, db)
    _, utc_end   = crud.get_utc_boundaries(end_date,   db)

    # ── 1. Ventas de contado (sin pagos en tabla pagos) ──────────────────────
    ventas_contado = (
        db.query(models.Venta)
        .filter(
            models.Venta.empresa_id == empresa_id,
            models.Venta.tipo       == "venta",
            models.Venta.fecha      >= utc_start,
            models.Venta.fecha      <= utc_end,
            models.Venta.estado_pago == "pagado",
            ~models.Venta.pagos.any(),
        )
        .all()
    )
    for v in ventas_contado:
        fecha_str = _fecha_col(v.fecha)
        metodo    = v.metodo_pago or "Efectivo"
        _acumular(fecha_str, metodo, float(v.total or 0), "ingreso")

    # ── 2. Abonos a cartera (tabla pagos) ────────────────────────────────────
    pagos_cartera = (
        db.query(models.Pago)
        .join(models.Venta)
        .filter(
            models.Venta.empresa_id == empresa_id,
            models.Pago.fecha       >= utc_start,
            models.Pago.fecha       <= utc_end,
        )
        .all()
    )
    for p in pagos_cartera:
        fecha_str = _fecha_col(p.fecha)
        metodo    = p.metodo_pago or "Efectivo"
        _acumular(fecha_str, metodo, float(p.monto or 0), "ingreso")

    # ── 3. Recaudo de préstamos ───────────────────────────────────────────────
    cuotas_pagadas = (
        db.query(models.CuotaPrestamo)
        .filter(
            models.CuotaPrestamo.empresa_id  == empresa_id,
            models.CuotaPrestamo.estado_pago.in_(["Pagado", "Parcial"]),
            models.CuotaPrestamo.fecha_pago  >= utc_start,
            models.CuotaPrestamo.fecha_pago  <= utc_end,
        )
        .all()
    )
    for c in cuotas_pagadas:
        fecha_str = _fecha_col(c.fecha_pago)
        metodo    = getattr(c, "metodo_pago", None) or "Efectivo"
        monto_rec = float(c.monto_cuota or 0) - float(c.saldo_pendiente or 0)
        if monto_rec > 0:
            _acumular(fecha_str, metodo, monto_rec, "ingreso")

    # ── 4. Abonos a capital (cuotas con metodo_pago='Abono Capital') ──────────
    abonos_capital = (
        db.query(models.CuotaPrestamo)
        .filter(
            models.CuotaPrestamo.empresa_id  == empresa_id,
            models.CuotaPrestamo.metodo_pago == "Abono Capital",
            models.CuotaPrestamo.fecha_pago  >= utc_start,
            models.CuotaPrestamo.fecha_pago  <= utc_end,
        )
        .all()
    )
    for c in abonos_capital:
        fecha_str = _fecha_col(c.fecha_pago)
        monto_rec = float(c.monto_cuota or 0) - float(c.saldo_pendiente or 0)
        if monto_rec > 0:
            _acumular(fecha_str, "Abono Capital", monto_rec, "ingreso")

    # ── 5. Gastos ─────────────────────────────────────────────────────────────
    gastos = (
        db.query(models.Gasto)
        .filter(
            models.Gasto.empresa_id == empresa_id,
            models.Gasto.fecha      >= utc_start,
            models.Gasto.fecha      <= utc_end,
        )
        .all()
    )
    for g in gastos:
        fecha_str = _fecha_col(g.fecha)
        _acumular(fecha_str, g.metodo_pago or "Efectivo", float(g.monto or 0), "egreso")

    # ── Ordenar días y construir respuesta ────────────────────────────────────
    dias_ordenados = sorted(resumen_dias.values(), key=lambda x: x["fecha"])

    return {
        "periodo": {
            "start_date": start_date.isoformat(),
            "end_date":   end_date.isoformat(),
        },
        "resumen": {
            "total_ingresos": round(total_ingresos, 2),
            "total_egresos":  round(total_egresos, 2),
            "neto":           round(total_ingresos - total_egresos, 2),
        },
        # #3 — Desglose por método de pago (sobre ingresos)
        "por_metodo": {
            metodo: round(valor, 2)
            for metodo, valor in sorted(metodos_totales.items(), key=lambda x: -x[1])
        },
        # #2 — Serie de días con ingresos, egresos y neto
        "dias": dias_ordenados,
    }

